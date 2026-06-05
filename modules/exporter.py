import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm

_RED    = "FFEF4444"
_AMBER  = "FFF59E0B"
_GREEN  = "FF22C55E"
_NAVY   = "FF0F1923"
_WHITE  = "FFFFFFFF"
_LGRAY  = "FFF1F5F9"


def _risk_fill(level: str) -> PatternFill:
    color = {"High": _RED, "Medium": _AMBER, "Low": _GREEN}.get(level, _LGRAY)
    return PatternFill("solid", fgColor=color)


def _header_fill() -> PatternFill:
    return PatternFill("solid", fgColor=_NAVY)


def _write_sheet(ws, df: pd.DataFrame, title: str):
    ws.title = title
    headers = list(df.columns)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _header_fill()
        cell.font = Font(color=_WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row in df.itertuples(index=False):
        ws.append(list(row))

    # Auto-width
    for i, col in enumerate(headers, 1):
        col_letter = get_column_letter(i)
        max_len = max(len(str(col)), df[col].astype(str).str.len().max() if len(df) > 0 else 10)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # Color risk_level column if present
    if "risk_level" in headers:
        ri = headers.index("risk_level") + 1
        for row_num in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=ri)
            cell.fill = _risk_fill(cell.value)
            cell.font = Font(bold=True)


def to_excel(scored_df: pd.DataFrame, provider_df: pd.DataFrame,
             member_df: pd.DataFrame, cost_df: pd.DataFrame) -> bytes:
    wb = Workbook()

    # Sheet 1: Flagged solicitações
    ws1 = wb.active
    flagged_cols = [c for c in ["claim_id", "member_id", "provider_id", "claim_date",
                                 "claim_amount", "risk_score", "risk_level", "risk_flags"]
                    if c in scored_df.columns]
    flagged = scored_df[scored_df["risk_level"].isin(["High", "Medium"])][flagged_cols].sort_values(
        "risk_score", ascending=False)
    _write_sheet(ws1, flagged.head(500), "Solicitações Sinalizadas")

    # Sheet 2: All solicitações
    ws2 = wb.create_sheet("Todas as Solicitações")
    all_cols = [c for c in flagged_cols if c in scored_df.columns]
    _write_sheet(ws2, scored_df[all_cols].sort_values("risk_score", ascending=False).head(2000), "Todas as Solicitações")

    # Sheet 3: Provider Risk
    ws3 = wb.create_sheet("Risco de Prestadores")
    p_cols = [c for c in ["provider_id", "provider_risk_score", "solicitação_count", "avg_amount",
                           "dup_rate", "round_rate", "provider_flags"] if c in provider_df.columns]
    _write_sheet(ws3, provider_df[p_cols].sort_values("provider_risk_score", ascending=False), "Risco de Prestadores")

    # Sheet 4: Member Risk
    ws4 = wb.create_sheet("Risco de Beneficiários")
    m_cols = [c for c in ["member_id", "member_risk_score", "solicitação_count", "total_spend",
                           "distinct_providers", "member_flags"] if c in member_df.columns]
    _write_sheet(ws4, member_df[m_cols].sort_values("member_risk_score", ascending=False), "Risco de Beneficiários")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_csv(scored_df: pd.DataFrame) -> bytes:
    cols = [c for c in ["claim_id", "member_id", "provider_id", "claim_date", "claim_amount",
                         "risk_score", "risk_level", "risk_flags"] if c in scored_df.columns]
    return scored_df[cols].sort_values("risk_score", ascending=False).to_csv(index=False).encode()


def to_pdf(scored_df: pd.DataFrame, provider_df: pd.DataFrame, member_df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle("title", fontSize=18, textColor=colors.HexColor("#0F1923"),
                                  spaceAfter=12, fontName="Helvetica-Bold")
    sub_style = ParagraphStyle("sub", fontSize=11, textColor=colors.HexColor("#1E2D3D"),
                                spaceAfter=8, fontName="Helvetica-Bold")
    body_style = styles["Normal"]

    story.append(Paragraph("Health Solicitações Intelligence Platform", title_style))
    story.append(Paragraph("Fraud & Anomaly Investigation Report", sub_style))
    story.append(Spacer(1, 0.5*cm))

    # Summary stats
    total = len(scored_df)
    high = (scored_df["risk_level"] == "High").sum() if "risk_level" in scored_df.columns else 0
    med  = (scored_df["risk_level"] == "Medium").sum() if "risk_level" in scored_df.columns else 0
    total_amt = pd.to_numeric(scored_df.get("claim_amount", pd.Series()), errors="coerce").sum()

    summary_data = [
        ["Metric", "Value"],
        ["Total de Solicitações Analisadas", f"{total:,}"],
        ["Solicitações de Alto Risco", f"{high:,} ({high/total*100:.1f}%)"],
        ["Solicitações de Risco Médio", f"{med:,} ({med/total*100:.1f}%)"],
        ["Valor Total Facturado", f"${total_amt:,.2f}"],
        ["Prestadores de Alto Risco", str((provider_df["provider_risk_score"] >= 70).sum()) if len(provider_df) > 0 else "N/A"],
    ]

    t = Table(summary_data, colWidths=[8*cm, 8*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F1923")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#F8FAFC"), colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.8*cm))

    # Top 20 flagged solicitações
    story.append(Paragraph("Top Flagged Solicitações", sub_style))
    flagged = scored_df[scored_df["risk_level"] == "High"].nlargest(20, "risk_score") if "risk_level" in scored_df.columns else scored_df.head(20)

    solicitação_cols = [c for c in ["claim_id", "provider_id", "claim_amount", "risk_score", "risk_flags"] if c in flagged.columns]
    headers = [c.replace("_", " ").title() for c in solicitação_cols]
    rows = [headers] + [[str(row[c])[:40] for c in solicitação_cols] for _, row in flagged.iterrows()]

    col_widths = [3*cm, 3*cm, 2.5*cm, 2*cm, 7*cm][:len(solicitação_cols)]
    ct = Table(rows, colWidths=col_widths, repeatRows=1)
    ct.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F1923")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#FEF2F2"), colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
        ("PADDING", (0, 0), (-1, -1), 4),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(ct)

    doc.build(story)
    return buf.getvalue()
