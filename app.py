import logging
import os

# Suppress harmless Windows WinError 10054 noise from asyncio ProactorEventLoop
logging.getLogger("asyncio").setLevel(logging.CRITICAL)

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime

from modules.column_detector import detect_columns
from modules.data_loader import (parse_upload, save_session, list_sessions,
                                  load_session, delete_session,
                                  save_feedback, load_feedback, get_feedback_stats,
                                  save_evaluation, load_evaluations, DB_PATH)
from modules import fraud_detection, provider_risk, member_utilization, cost_outlier, risk_scorer, exporter
from modules import supervised_model
from modules.lang import t, translate_flag

# ── Language init (before any st calls) ───────────────────────────────────────
if "lang" not in st.session_state:
    st.session_state.lang = "pt"

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Health Insurance Fraud Detection Platform",
    page_icon="🔍",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Theme ─────────────────────────────────────────────────────────────────────
THEME_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');

html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.stApp { background-color: #080F18; color: #E2E8F0; }

/* ── Sidebar ── */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0D1B2A 0%, #0A1520 100%) !important;
    border-right: 1px solid #1E3A50;
}
[data-testid="stSidebar"] * { color: #E2E8F0 !important; }

/* ── Radio nav ── */
[data-testid="stSidebar"] .stRadio > label { display: none !important; }
[data-testid="stSidebar"] .stRadio > div[role="radiogroup"] {
    display: flex !important; flex-direction: column !important; gap: 3px !important;
}
[data-testid="stSidebar"] .stRadio > div[role="radiogroup"] > label {
    display: flex !important; align-items: center !important;
    padding: 0.6rem 1.1rem !important; border-radius: 8px !important;
    cursor: pointer !important; transition: all 0.2s ease !important;
    font-size: 0.88rem !important; font-weight: 500 !important;
    color: #64748B !important; border: none !important;
    background: transparent !important; width: 100% !important;
    letter-spacing: 0.01em !important;
}
[data-testid="stSidebar"] .stRadio > div[role="radiogroup"] > label:hover {
    background: rgba(255,255,255,0.05) !important; color: #CBD5E1 !important;
}
[data-testid="stSidebar"] .stRadio > div[role="radiogroup"] > label:has(input:checked) {
    background: rgba(245,158,11,0.1) !important;
    border-left: 3px solid #F59E0B !important;
    color: #F59E0B !important; font-weight: 700 !important;
    padding-left: calc(1.1rem - 3px) !important;
}
[data-testid="stSidebar"] .stRadio > div[role="radiogroup"] > label > div:first-child { display: none !important; }
[data-testid="stSidebar"] .stRadio > div[role="radiogroup"] > label > div:last-child {
    color: inherit !important; font-size: inherit !important; font-weight: inherit !important;
}

/* ── Main content area ── */
.main .block-container { padding: 1.5rem 2rem 3rem 2rem !important; max-width: 100% !important; }

/* ── Section headers ── */
h1, h2, h3 { color: #F1F5F9 !important; letter-spacing: -0.3px; }
h2 { border-bottom: 1px solid #1E3A50; padding-bottom: 0.5rem; margin-bottom: 1rem !important; }

/* ── KPI cards ── */
.card {
    background: linear-gradient(135deg, #112233 0%, #0D1B2A 100%);
    border: 1px solid #1E3A50;
    border-radius: 12px;
    padding: 1.3rem 1.4rem;
    margin-bottom: 0.8rem;
    transition: border-color 0.2s ease;
}
.card:hover { border-color: #2D5A7A; }
.kpi-value { font-size: 2.1rem; font-weight: 800; line-height: 1.1; letter-spacing: -0.5px; }
.kpi-label { font-size: 0.72rem; color: #475569; text-transform: uppercase; letter-spacing: 0.08em; margin-top: 4px; }
.kpi-high   { color: #F87171 !important; }
.kpi-medium { color: #FBBF24 !important; }
.kpi-low    { color: #34D399 !important; }
.kpi-blue   { color: #60A5FA !important; }
.kpi-purple { color: #A78BFA !important; }

/* ── st.metric ── */
[data-testid="stMetric"] {
    background: linear-gradient(135deg, #112233 0%, #0D1B2A 100%);
    border: 1px solid #1E3A50; border-radius: 12px; padding: 1rem 1.2rem;
}
[data-testid="stMetricLabel"] { color: #475569 !important; font-size: 0.72rem !important; text-transform: uppercase; letter-spacing: 0.07em; }
[data-testid="stMetricValue"] { color: #F1F5F9 !important; font-size: 1.7rem !important; font-weight: 800 !important; letter-spacing: -0.5px; }
[data-testid="stMetricDelta"] { color: #64748B !important; }

/* ── Tables ── */
.dataframe { border-radius: 8px; overflow: hidden; }
.dataframe thead th { background-color: #112233 !important; color: #94A3B8 !important; font-size: 0.75rem !important; text-transform: uppercase; letter-spacing: 0.06em; border-bottom: 1px solid #1E3A50 !important; }
.dataframe tbody tr:nth-child(even) { background-color: #0D1B2A !important; }
.dataframe tbody tr:nth-child(odd)  { background-color: #080F18 !important; }
.dataframe tbody tr:hover { background-color: #112233 !important; }

/* ── Inputs ── */
.stSelectbox label, .stSlider label, .stFileUploader label { color: #64748B !important; font-size: 0.8rem !important; text-transform: uppercase; letter-spacing: 0.06em; }
[data-testid="stFileUploader"] { background: #0D1B2A; border: 2px dashed #1E3A50; border-radius: 10px; }

/* ── Alert cards ── */
.alert-high   { background:linear-gradient(135deg,#2D1515,#1A0A0A); border-left:4px solid #F87171; padding:0.9rem 1.1rem; border-radius:8px; margin:6px 0; }
.alert-medium { background:linear-gradient(135deg,#2D2415,#1A1500); border-left:4px solid #FBBF24; padding:0.9rem 1.1rem; border-radius:8px; margin:6px 0; }
.alert-low    { background:linear-gradient(135deg,#152D1A,#0A1A0F); border-left:4px solid #34D399; padding:0.9rem 1.1rem; border-radius:8px; margin:6px 0; }

/* ── Adjudication badges ── */
.badge-investigate { background:#2D1515; color:#F87171; border:1px solid #F8717140; padding:3px 10px; border-radius:20px; font-size:0.72rem; font-weight:700; }
.badge-review      { background:#2D2415; color:#FBBF24; border:1px solid #FBBF2440; padding:3px 10px; border-radius:20px; font-size:0.72rem; font-weight:700; }
.badge-approve     { background:#152D1A; color:#34D399; border:1px solid #34D39940; padding:3px 10px; border-radius:20px; font-size:0.72rem; font-weight:700; }

/* ── Tabs ── */
[data-testid="stTabs"] button { color: #64748B !important; font-weight: 600; }
[data-testid="stTabs"] button[aria-selected="true"] { color: #F59E0B !important; border-bottom-color: #F59E0B !important; }

/* ── Scrollbar ── */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #0D1B2A; }
::-webkit-scrollbar-thumb { background: #1E3A50; border-radius: 3px; }

/* ── Form labels — force white text ── */
[data-testid="stForm"] label,
[data-testid="stForm"] .stRadio label,
[data-testid="stForm"] .stSelectbox label,
[data-testid="stForm"] .stTextInput label,
[data-testid="stForm"] .stTextArea label,
[data-testid="stForm"] .stSlider label,
[data-testid="stForm"] p,
[data-testid="stForm"] span { color: #F1F5F9 !important; }
[data-testid="stForm"] .stRadio > div[role="radiogroup"] > label {
    color: #CBD5E1 !important;
    background: rgba(255,255,255,0.03) !important;
    border: 1px solid #1E3A50 !important;
    border-radius: 6px !important;
    padding: 0.4rem 0.8rem !important;
    margin-bottom: 4px !important;
}
[data-testid="stForm"] .stRadio > div[role="radiogroup"] > label:has(input:checked) {
    background: rgba(245,158,11,0.12) !important;
    border-color: #F59E0B !important;
    color: #F59E0B !important;
}
[data-testid="stForm"] .stSlider [data-testid="stTickBar"] span { color: #64748B !important; }
[data-testid="stForm"] input,
[data-testid="stForm"] textarea {
    background: #0D1B2A !important;
    color: #E2E8F0 !important;
    border: 1px solid #1E3A50 !important;
}
[data-testid="stForm"] .stSelectbox > div > div {
    background: #0D1B2A !important;
    color: #E2E8F0 !important;
    border: 1px solid #1E3A50 !important;
}
</style>
"""
st.markdown(THEME_CSS, unsafe_allow_html=True)

# ── Session state ──────────────────────────────────────────────────────────────
if "scored_df" not in st.session_state:
    st.session_state.scored_df = None
if "provider_df" not in st.session_state:
    st.session_state.provider_df = None
if "member_df" not in st.session_state:
    st.session_state.member_df = None
if "profile" not in st.session_state:
    st.session_state.profile = None
if "active_session" not in st.session_state:
    st.session_state.active_session = None
if "trends_df" not in st.session_state:
    st.session_state.trends_df = None
if "session_id" not in st.session_state:
    st.session_state.session_id = "demo"
if "supervised_metrics" not in st.session_state:
    st.session_state.supervised_metrics = supervised_model.load_metrics()
if "model_pkl_mtime" not in st.session_state:
    # Track pickle modification time to detect external model updates
    _pkl = supervised_model.MODEL_PATH
    st.session_state.model_pkl_mtime = _pkl.stat().st_mtime if _pkl.exists() else 0


# ── Sidebar navigation ─────────────────────────────────────────────────────────
def build_nav():
    return [
        f"📊  {t('nav_overview')}",
        f"🔎  {t('nav_claims')}",
        f"🏥  {t('nav_providers')}",
        f"👤  {t('nav_members')}",
        f"💰  {t('nav_costs')}",
        f"📋  {t('nav_report')}",
        f"📁  {t('nav_data')}",
        f"ℹ️  {t('nav_howto')}",
        f"⭐  {t('nav_eval')}",
    ]

with st.sidebar:
    # ── Language toggle ────────────────────────────────────────────────────────
    st.markdown(
        "<div style='padding:0.6rem 0.5rem 0.2rem 0.5rem;"
        "font-size:0.7rem;color:#475569;text-transform:uppercase;"
        "letter-spacing:0.08em'>Idioma / Language</div>",
        unsafe_allow_html=True,
    )
    lc1, lc2 = st.columns(2)
    with lc1:
        if st.button("PT", use_container_width=True,
                     type="primary" if st.session_state.lang == "pt" else "secondary"):
            st.session_state.lang = "pt"
            st.rerun()
    with lc2:
        if st.button("EN", use_container_width=True,
                     type="primary" if st.session_state.lang == "en" else "secondary"):
            st.session_state.lang = "en"
            st.rerun()

    st.markdown("<hr style='border-color:#1E3A50;margin:0.5rem 0 0.4rem 0'>", unsafe_allow_html=True)

    st.markdown(
        f"<div style='padding:0.6rem 0.5rem 0.6rem 0.8rem'>"
        f"<span style='font-size:1.05rem;font-weight:800;color:#F59E0B;letter-spacing:-0.3px'>"
        f"{t('sidebar_title')}</span></div>",
        unsafe_allow_html=True,
    )
    st.markdown("<hr style='border-color:#1E3A50;margin:0 0 0.4rem 0'>", unsafe_allow_html=True)

    NAV_OPTIONS = build_nav()
    nav_selection = st.radio("nav", NAV_OPTIONS, label_visibility="hidden")

    st.markdown("<hr style='border-color:#1E3A50;margin:0.4rem 0'>", unsafe_allow_html=True)

    if st.session_state.scored_df is not None:
        _df = st.session_state.scored_df
        _high = (_df["risk_level"] == "High").sum()
        st.markdown(
            f"<div style='padding:0.7rem 0.8rem;background:#0D1B2A;"
            f"border-radius:8px;border:1px solid #1E3A50'>"
            f"<div style='color:#475569;font-size:0.7rem;text-transform:uppercase;"
            f"letter-spacing:0.06em;margin-bottom:4px'>{t('sidebar_session')}</div>"
            f"<div style='color:#E2E8F0;font-weight:600;font-size:0.95rem'>"
            f"{len(_df):,} {t('sidebar_claims')}</div>"
            f"<div style='color:#F87171;font-size:0.82rem;margin-top:2px'>"
            f"&#11044; {_high} {t('sidebar_highrisk')}</div>"
            f"</div>",
            unsafe_allow_html=True,
        )
        if st.session_state.profile and st.session_state.profile.missing_features:
            with st.expander(f"ℹ️ {t('sidebar_skipped')}"):
                for note in st.session_state.profile.missing_features:
                    st.caption(f"• {note}")

page = nav_selection.split("  ", 1)[1] if "  " in nav_selection else nav_selection


# ── Analysis pipeline ──────────────────────────────────────────────────────────
def run_analysis(df_raw: pd.DataFrame):
    with st.spinner(t("spin_columns")):
        df, profile, missing = detect_columns(df_raw)

    if missing:
        st.error(f"Colunas obrigatórias em falta: {', '.join(missing)}. Por favor, adicione-as ao ficheiro.")
        st.stop()

    df["claim_amount"] = pd.to_numeric(df["claim_amount"], errors="coerce").fillna(0)
    if "claim_date" in df.columns:
        df["claim_date"] = pd.to_datetime(df["claim_date"], errors="coerce")

    with st.spinner(t("spin_anomaly")):
        anomaly_res = fraud_detection.run(df, profile)

    with st.spinner(t("spin_providers")):
        prov_claim_scores, provider_df = provider_risk.run(df, profile)

    with st.spinner(t("spin_members")):
        mem_claim_scores, member_df = member_utilization.run(df, profile)

    with st.spinner(t("spin_costs")):
        cost_res = cost_outlier.run(df, profile)

    with st.spinner(t("spin_scoring")):
        scored = risk_scorer.compute(df, anomaly_res, prov_claim_scores, mem_claim_scores, cost_res)

    trends_df = provider_risk.monthly_trends(df)

    # Stage 2: apply supervised classifier if already trained
    with st.spinner("Aplicando modelo supervisionado..." if st.session_state.lang == "pt" else "Applying supervised model..."):
        scored = supervised_model.predict(scored)

    st.session_state.scored_df   = scored
    st.session_state.provider_df = provider_df
    st.session_state.member_df   = member_df
    st.session_state.profile     = profile
    st.session_state.trends_df   = trends_df
    st.session_state.session_id  = save_session(df, getattr(df, "_filename", "demo"))
    st.success(t("analysis_complete"))


# ──────────────────────────────────────────────────────────────────────────────
# PÁGINA: Gestão de Dados
# ──────────────────────────────────────────────────────────────────────────────
if page == t("nav_data"):
    st.title(t("data_title"))

    # ── Dados de demonstração ──────────────────────────────────────────────────
    st.markdown(
        f'<div style="background:linear-gradient(135deg,#1E3A2F 0%,#162030 100%);'
        f'border:1px solid #22C55E40;border-radius:10px;padding:1rem 1.4rem;margin-bottom:1.2rem">'
        f'<div style="font-size:1rem;font-weight:700;color:#22C55E;margin-bottom:0.3rem">'
        f'🧪 {t("data_demo_title")}</div>'
        f'<div style="font-size:0.85rem;color:#94A3B8">{t("data_demo_desc")}</div>'
        f'</div>',
        unsafe_allow_html=True,
    )

    if st.button(t("data_demo_btn"), type="primary", use_container_width=True):
        import random
        rng = np.random.default_rng(42)
        random.seed(42)
        n_normal = 1800
        specialties = ["Cardiologia","Ortopedia","Clínica Geral","Oncologia",
                        "Neurologia","Radiologia","Fisioterapia","Psiquiatria"]
        proc_codes  = ["99213","99214","99215","27447","93000","71046",
                        "99232","90837","20610","70553","99283","43239"]
        diag_codes  = ["Z00.00","I10","M54.5","E11.9","J06.9","K21.0",
                        "F32.9","M17.11","G89.29","Z12.31","I25.10","N39.0"]
        providers   = [f"PRV{str(i).zfill(4)}" for i in range(1, 41)]
        fraud_provs = providers[:3]
        members     = [f"MBR{str(i).zfill(5)}" for i in range(1, 301)]
        fraud_mems  = members[:5]
        prov_spec   = {p: random.choice(specialties) for p in providers}
        mem_ages    = {m: random.randint(18, 80) for m in members}
        mem_genders = {m: random.choice(["M","F"]) for m in members}

        import datetime as _dt
        start = _dt.date(2023, 1, 1)
        def rdate(): return start + _dt.timedelta(days=int(rng.integers(0, 730)))
        def ramt(mn=400, sd=250): return round(float(np.clip(rng.normal(mn, sd), 50, 8000)), 2)

        rows = []
        for i in range(n_normal):
            cd = rdate(); p = random.choice(providers); m = random.choice(members); a = ramt()
            rows.append({"claim_id": f"CLM{i+1:06d}", "member_id": m, "provider_id": p,
                          "claim_date": str(cd), "service_date": str(cd - _dt.timedelta(days=random.randint(0,5))),
                          "claim_amount": a, "paid_amount": round(a*random.uniform(0.7,0.95),2),
                          "diagnosis_code": random.choice(diag_codes),
                          "procedure_code": random.choice(proc_codes),
                          "provider_specialty": prov_spec[p],
                          "member_age": mem_ages[m], "member_gender": mem_genders[m],
                          "claim_type": random.choice(["Médico","Farmácia","Dentário"])})
        # Outliers
        for i in range(30):
            p = random.choice(fraud_provs); cd = rdate(); a = round(random.uniform(8000,25000),2)
            rows.append({"claim_id": f"FRD{i+1:06d}", "member_id": random.choice(members),
                          "provider_id": p, "claim_date": str(cd), "service_date": str(cd),
                          "claim_amount": a, "paid_amount": round(a*0.9,2),
                          "diagnosis_code": "I10", "procedure_code": "99215",
                          "provider_specialty": prov_spec[p],
                          "member_age": 45, "member_gender": "M", "claim_type": "Médico"})
        # Duplicados
        for dup in rows[:10]:
            r = dup.copy(); r["claim_id"] = f"DUP{r['claim_id']}"; rows.append(r)
        # Valores redondos
        for i in range(25):
            p = random.choice(fraud_provs); cd = rdate(); a = float(random.choice([500,1000,1500,2000,5000]))
            rows.append({"claim_id": f"RND{i+1:06d}", "member_id": random.choice(members),
                          "provider_id": p, "claim_date": str(cd), "service_date": str(cd),
                          "claim_amount": a, "paid_amount": a*0.85, "diagnosis_code": random.choice(diag_codes),
                          "procedure_code": "99215", "provider_specialty": prov_spec[p],
                          "member_age": 40, "member_gender": "F", "claim_type": "Médico"})
        # Multi-prestadores
        many_provs = providers[5:25]
        for mem in fraud_mems:
            for j in range(12):
                p = many_provs[j % len(many_provs)]; cd = rdate(); a = ramt(600,150)
                rows.append({"claim_id": f"SHP{mem[-5:]}{j:03d}", "member_id": mem,
                              "provider_id": p, "claim_date": str(cd), "service_date": str(cd),
                              "claim_amount": a, "paid_amount": round(a*0.88,2),
                              "diagnosis_code": random.choice(diag_codes),
                              "procedure_code": random.choice(proc_codes),
                              "provider_specialty": prov_spec[p],
                              "member_age": mem_ages[mem], "member_gender": mem_genders[mem],
                              "claim_type": "Médico"})

        df_demo = pd.DataFrame(rows).sample(frac=1, random_state=42).reset_index(drop=True)
        sid = save_session(df_demo, "dados_demonstracao.csv")
        st.session_state.active_session = sid
        run_analysis(df_demo)
        st.rerun()

    st.markdown("<hr style='border-color:#2D3F50;margin:1rem 0'>", unsafe_allow_html=True)

    col1, col2 = st.columns([1, 1])

    with col1:
        st.subheader(t("data_upload_title"))
        uploaded = st.file_uploader("CSV ou Excel (.xlsx)", type=["csv", "xlsx"])
        contamination = st.slider("Sensibilidade a anomalias (% esperada de outliers)", 1, 20, 5) / 100

        if uploaded:
            if st.button(t("data_analyze_btn"), type="primary"):
                df_raw = parse_upload(uploaded)
                # Memory guard: Streamlit Cloud free tier has ~1 GB RAM and the
                # pipeline keeps several scored copies of the DataFrame in memory
                MAX_ROWS = 50_000
                if len(df_raw) > MAX_ROWS:
                    st.error(
                        f"⚠️ O ficheiro tem {len(df_raw):,} linhas — o limite "
                        f"suportado é {MAX_ROWS:,}. Divida o ficheiro por mês "
                        f"ou por período mais curto e carregue novamente."
                    )
                    st.stop()
                sid = save_session(df_raw, uploaded.name)
                st.session_state.active_session = sid
                run_analysis(df_raw)
                st.rerun()

    with col2:
        st.subheader(t("data_sessions"))
        sessions = list_sessions()
        if len(sessions) > 0:
            for _, row in sessions.iterrows():
                c1, c2, c3 = st.columns([3, 1, 1])
                with c1:
                    st.markdown(f"**{row['filename']}**  \n{row['uploaded_at'][:16]} · {row['row_count']:,} {t('data_records')}")
                with c2:
                    if st.button(t("data_load_btn"), key=f"load_{row['session_id']}"):
                        df_raw = load_session(row["session_id"])
                        st.session_state.active_session = row["session_id"]
                        run_analysis(df_raw)
                        st.rerun()
                with c3:
                    if st.button("🗑️", key=f"del_{row['session_id']}"):
                        delete_session(row["session_id"])
                        st.rerun()
                st.divider()
        else:
            st.info("Nenhuma sessão anterior encontrada. Carregue um ficheiro para começar.")

    # ── Backup / Restore da base de dados ────────────────────────────────────
    # Streamlit Cloud has an ephemeral filesystem: claims.db is wiped on every
    # redeploy or container restart. These buttons let the team export the DB
    # (sessions + investigator feedback + evaluations) and restore it afterwards.
    st.subheader("💾 Backup e Restauro da Base de Dados")
    _bk_col1, _bk_col2 = st.columns([1, 1])

    with _bk_col1:
        if DB_PATH.exists():
            with open(DB_PATH, "rb") as _dbf:
                st.download_button(
                    label="⬇️ Descarregar Backup (.db)",
                    data=_dbf.read(),
                    file_name=f"claims_backup_{datetime.now():%Y%m%d_%H%M}.db",
                    mime="application/octet-stream",
                    use_container_width=True,
                    help="Guarda sessões, feedback dos investigadores e avaliações. "
                         "Faça backup semanal — os dados são apagados em cada redeploy.",
                )
        else:
            st.info("Ainda não existe base de dados para exportar.")

    with _bk_col2:
        _restore_file = st.file_uploader(
            "Restaurar Backup (.db)", type=["db"], key="db_restore_upload"
        )
        if _restore_file is not None:
            if st.button("⚠️ Confirmar Restauro (substitui a BD actual)",
                         use_container_width=True):
                DB_PATH.parent.mkdir(parents=True, exist_ok=True)
                # Atomic replace so a failed upload can't corrupt the live DB
                _tmp_db = DB_PATH.with_suffix(".db.tmp")
                with open(_tmp_db, "wb") as _out:
                    _out.write(_restore_file.getbuffer())
                os.replace(_tmp_db, DB_PATH)
                st.success("✅ Base de dados restaurada. A recarregar...")
                st.rerun()

    st.markdown("<hr style='border-color:#2D3F50;margin:1rem 0'>", unsafe_allow_html=True)

    # ── Model Administration (technical — lives here, not on claims page) ────
    st.subheader("🤖 Estado do Modelo de IA")
    _sm = st.session_state.get("supervised_metrics")
    _fb_admin = load_feedback(st.session_state.get("session_id", "demo"))

    _adm_col1, _adm_col2 = st.columns([3, 2])
    with _adm_col1:
        if _sm and _sm.get("status") == "trained":
            st.markdown(
                f'<div style="background:#1a2d1a;border-left:4px solid #22C55E;border-radius:6px;'
                f'padding:0.8rem 1rem">'
                f'<strong style="color:#22C55E">✅ Modelo supervisionado activo</strong><br>'
                f'<span style="color:#94A3B8">Treinado com '
                f'<strong style="color:#E2E8F0">{_sm.get("n_labeled",0)}</strong> exemplos · '
                f'<strong style="color:#EF4444">{_sm.get("n_fraud",0)} fraude</strong> · '
                f'<strong style="color:#22C55E">{_sm.get("n_legit",0)} legítimos</strong></span><br>'
                f'<span style="color:#64748B;font-size:0.82rem">'
                f'Precisão: {_sm.get("precision",0):.0%} · '
                f'Recall: {_sm.get("recall",0):.0%} · '
                f'F1: {_sm.get("f1",0):.0%}'
                + (f' · F1 CV: {_sm["f1_cv_mean"]:.0%}' if _sm.get("f1_cv_mean") else '')
                + f'</span></div>',
                unsafe_allow_html=True
            )
        elif _sm and _sm.get("status") == "insufficient":
            _nf = _sm.get("n_fraud", 0); _nl = _sm.get("n_legit", 0)
            st.markdown(
                f'<div style="background:#1a1a2d;border-left:4px solid #A78BFA;border-radius:6px;'
                f'padding:0.8rem 1rem">'
                f'<strong style="color:#A78BFA">⏳ Dados insuficientes</strong><br>'
                f'<span style="color:#94A3B8">Necessários ≥5 de cada classe · '
                f'Actuais: {_nf} fraude, {_nl} legítimos</span></div>',
                unsafe_allow_html=True
            )
        else:
            st.markdown(
                '<div style="background:#1a1a2d;border-left:4px solid #64748B;border-radius:6px;'
                'padding:0.8rem 1rem">'
                '<strong style="color:#94A3B8">🔒 Modelo não treinado</strong><br>'
                '<span style="color:#64748B">Marque claims na página Análise de Solicitações '
                'para acumular dados de treino.</span></div>',
                unsafe_allow_html=True
            )

    with _adm_col2:
        if st.button("🔄 Forçar Re-treino do Modelo", use_container_width=True):
            with st.spinner("A treinar..."):
                _m, _metrics = supervised_model.train(
                    st.session_state.scored_df, _fb_admin
                )
            if _metrics and _metrics.get("status") == "trained":
                st.session_state.scored_df = supervised_model.predict(
                    st.session_state.scored_df, _m
                )
                st.session_state.supervised_metrics = _metrics
                st.success(f"✅ Re-treinado com {_metrics['n_labeled']} exemplos.")
                st.rerun()
            else:
                msg = (_metrics or {}).get("message", "Dados insuficientes.")
                st.warning(f"⚠️ {msg}")
                st.session_state.supervised_metrics = _metrics

    # Feature importance chart (technical detail — admin only)
    if _sm and _sm.get("status") == "trained":
        _fi = _sm.get("feature_importance", {})
        if _fi:
            with st.expander("📊 Ver importância das features do modelo"):
                _fi_df = pd.DataFrame(list(_fi.items()), columns=["Feature", "Importância"])
                _fi_df = _fi_df.sort_values("Importância", ascending=True)
                _fi_colors = ["#F59E0B" if v > 0.2 else "#3B82F6" if v > 0.1 else "#64748B"
                              for v in _fi_df["Importância"]]
                _fig_fi = go.Figure(go.Bar(
                    x=_fi_df["Importância"], y=_fi_df["Feature"], orientation="h",
                    marker=dict(color=_fi_colors),
                    text=[f"{v:.1%}" for v in _fi_df["Importância"]],
                    textposition="outside", textfont=dict(color="#E2E8F0", size=10),
                    hovertemplate="<b>%{y}</b><br>Importância: %{x:.1%}<extra></extra>",
                ))
                _fig_fi.update_layout(
                    paper_bgcolor="#1E2D3D", plot_bgcolor="#1E2D3D",
                    font_color="#E2E8F0", height=280,
                    xaxis=dict(showgrid=True, gridcolor="#243447", tickformat=".0%",
                               tickfont=dict(color="#64748B")),
                    yaxis=dict(showgrid=False, tickfont=dict(color="#CBD5E1", size=10)),
                    margin=dict(t=10, b=20, l=180, r=60),
                )
                st.plotly_chart(_fig_fi, use_container_width=True)

    # Full feedback log (admin view)
    if not _fb_admin.empty:
        with st.expander(f"📋 Registo completo de feedback ({len(_fb_admin)} avaliações)"):
            st.dataframe(
                _fb_admin[["claim_id", "verdict", "investigator", "notes", "created_at"]],
                use_container_width=True
            )

    st.markdown("---")
    st.subheader(t("data_col_format"))
    _req = t("data_col_required")
    _opt = t("data_col_optional")
    if st.session_state.lang == "en":
        st.markdown(f"""
| Column | Required | Description |
|--------|----------|-------------|
| `claim_id` | ✅ | Unique claim identifier |
| `member_id` | ✅ | Member / patient identifier |
| `provider_id` | ✅ | Provider / facility identifier |
| `claim_date` | ✅ | Date claim was submitted (YYYY-MM-DD) |
| `claim_amount` | ✅ | Total billed amount |
| `diagnosis_code` | {_opt} | ICD-10 diagnosis code |
| `procedure_code` | {_opt} | CPT / HCPCS procedure code |
| `service_date` | {_opt} | Date of service |
| `paid_amount` | {_opt} | Amount approved / paid |
| `provider_specialty` | {_opt} | Provider specialty |
| `member_age` | {_opt} | Member age |
| `member_gender` | {_opt} | Member gender |
| `drug_name` / `ndc_code` | {_opt} | For pharmacy claims |
        """)
    else:
        st.markdown(f"""
| Coluna | {_req} | Descrição |
|--------|---------|-----------|
| `claim_id` | ✅ | Identificador único do claim |
| `member_id` | ✅ | Identificador do beneficiário / paciente |
| `provider_id` | ✅ | Identificador do prestador / estabelecimento |
| `claim_date` | ✅ | Data de submissão do claim (YYYY-MM-DD) |
| `claim_amount` | ✅ | Valor total facturado |
| `diagnosis_code` | {_opt} | Código de diagnóstico ICD-10 |
| `procedure_code` | {_opt} | Código de procedimento CPT / HCPCS |
| `service_date` | {_opt} | Data de prestação do serviço |
| `paid_amount` | {_opt} | Valor aprovado / pago |
| `provider_specialty` | {_opt} | Especialidade do prestador |
| `member_age` | {_opt} | Idade do beneficiário |
| `member_gender` | {_opt} | Género do beneficiário |
| `drug_name` / `ndc_code` | {_opt} | Para claims de farmácia |
        """)


# ── Como Funciona ─────────────────────────────────────────────────────────────
if page == t("nav_howto"):
    st.markdown("""
    <div style="background:linear-gradient(135deg,#1E2D3D 0%,#162030 100%);
                border:1px solid #2D3F50;border-left:5px solid #3B82F6;
                border-radius:12px;padding:1.4rem 1.8rem;margin-bottom:1.5rem">
        <div style="font-size:1.5rem;font-weight:800;color:#F1F5F9">
            ℹ️ Como Funciona a Plataforma
        </div>
        <div style="font-size:0.88rem;color:#94A3B8;margin-top:0.3rem">
            Guia completo de utilização e metodologia de detecção de fraude
        </div>
    </div>
    """, unsafe_allow_html=True)

    t1, t2, t3, t4 = st.tabs([
        t("howto_tab1"),
        t("howto_tab2"),
        t("howto_tab3"),
        t("howto_tab4"),
    ])

    with t1:
        st.markdown("### Como começar em 3 passos")
        steps = [
            ("1", "#3B82F6", "Carregar os dados",
             "Vá a **Gestão de Dados** no menu lateral e carregue um ficheiro CSV ou Excel com os seus solicitações. "
             "A plataforma detecta automaticamente as colunas disponíveis e adapta as análises ao que existe no ficheiro. "
             "Só são obrigatórias 5 colunas: `claim_id`, `member_id`, `provider_id`, `claim_date`, `claim_amount`."),
            ("2", "#F59E0B", "Analisar os resultados",
             "Após o carregamento, a plataforma executa automaticamente 4 módulos de análise em paralelo: "
             "detecção de anomalias, risco de prestadores, utilização de beneficiários e custos atípicos. "
             "Cada solicitação recebe uma **pontuação de risco de 0 a 100**."),
            ("3", "#22C55E", "Priorizar e exportar",
             "Use o **Painel de Controlo** para ver os alertas mais críticos, a **Análise de Solicitações** para filtrar e "
             "investigar casos específicos, e os botões de exportação para descarregar relatórios em Excel, CSV ou PDF "
             "para partilhar com a equipa de investigação."),
        ]
        for num, color, title, desc in steps:
            st.markdown(
                f'<div style="display:flex;gap:1rem;background:#1E2D3D;border:1px solid #2D3F50;'
                f'border-radius:10px;padding:1.1rem 1.3rem;margin-bottom:0.7rem;align-items:flex-start">'
                f'<div style="background:{color};color:#0F1923;font-size:1.1rem;font-weight:800;'
                f'border-radius:50%;width:36px;height:36px;display:flex;align-items:center;'
                f'justify-content:center;flex-shrink:0">{num}</div>'
                f'<div><div style="font-size:1rem;font-weight:700;color:#F1F5F9;margin-bottom:0.3rem">{title}</div>'
                f'<div style="font-size:0.87rem;color:#94A3B8;line-height:1.6">{desc}</div></div>'
                f'</div>',
                unsafe_allow_html=True,
            )

        st.markdown("---")
        st.markdown("### Níveis de Risco")
        for color, bg, border, level, desc in [
            ("#EF4444","#2D1515","#EF444440","🔴 Alto Risco (70–100)",
             "Solicitação com múltiplos sinais de fraude activos. Requer investigação prioritária. "
             "Inclui outliers estatísticos, duplicados e facturação anómala."),
            ("#F59E0B","#2D2415","#F59E0B40","🟡 Risco Médio (40–69)",
             "Solicitação com um ou mais sinais de atenção. Deve ser revisto pela equipa, "
             "mas pode ter justificação clínica ou administrativa legítima."),
            ("#22C55E","#152D1A","#22C55E40","🟢 Baixo Risco (0–39)",
             "Solicitação dentro dos padrões normais de facturação. Sem sinais de anomalia detectados "
             "nos modelos estatísticos e de comparação com pares."),
        ]:
            st.markdown(
                f'<div style="background:{bg};border:1px solid {border};border-radius:8px;'
                f'padding:0.8rem 1.1rem;margin-bottom:0.5rem">'
                f'<div style="font-weight:700;color:{color};margin-bottom:0.2rem">{level}</div>'
                f'<div style="font-size:0.85rem;color:#94A3B8">{desc}</div></div>',
                unsafe_allow_html=True,
            )

    with t2:
        st.markdown("### Fórmula de Pontuação de Risco Composta")
        st.markdown(
            '<div style="background:#1E2D3D;border:1px solid #2D3F50;border-radius:10px;'
            'padding:1.2rem 1.5rem;font-family:monospace;font-size:0.92rem;color:#A78BFA;margin-bottom:1rem">'
            'Pontuação Final = <br>'
            '&nbsp;&nbsp;Anomalia Estatística &times; 35% <br>'
            '&nbsp;&nbsp;+ Risco do Prestador &nbsp;&times; 25% <br>'
            '&nbsp;&nbsp;+ Risco do Beneficiário &times; 20% <br>'
            '&nbsp;&nbsp;+ Custo Atípico &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&times; 20%'
            '</div>',
            unsafe_allow_html=True,
        )

        st.markdown("### Sinais de Fraude Detectados")
        signals = [
            ("Isolation Forest", "Modelo de Machine Learning não-supervisionado que identifica solicitações "
             "estatisticamente isolados do comportamento normal da carteira. Detecta padrões complexos "
             "que as regras simples não conseguem capturar."),
            ("Z-Score por Prestador", "Compara o valor de cada solicitação com a média e desvio padrão dos "
             "claims do mesmo prestador. Um Z-Score > 3 indica que o valor está muito acima do habitual "
             "para aquele prestador específico."),
            ("Detecção de Duplicados", "Identifica solicitações com o mesmo beneficiário, prestador e valor "
             "na mesma data. Padrão comum em esquemas de facturação duplicada intencional."),
            ("Valores Redondos", "Valores exactamente redondos (ex: $500, $1.000, $2.500) são "
             "estatisticamente raros em facturação real e podem indicar facturação estimada ou fictícia."),
            ("Alta Frequência de Procedimentos", "Detecta prestadores que facturam determinados "
             "procedimentos com frequência muito superior à média dos pares — sinal clássico de upcoding."),
            ("Custo Atípico vs. Pares", "Compara o valor do solicitação com a média de outros solicitações "
             "semelhantes (mesmo procedimento, especialidade ou carteira geral). Valores > 2 desvios "
             "padrão acima da média são sinalizados."),
        ]
        for i, (title, desc) in enumerate(signals):
            color = ["#3B82F6","#F59E0B","#EF4444","#A78BFA","#22C55E","#F97316"][i % 6]
            st.markdown(
                f'<div style="background:#1E2D3D;border:1px solid #2D3F50;border-left:4px solid {color};'
                f'border-radius:8px;padding:0.9rem 1.1rem;margin-bottom:0.5rem">'
                f'<div style="font-weight:700;color:{color};margin-bottom:0.3rem">{title}</div>'
                f'<div style="font-size:0.85rem;color:#94A3B8;line-height:1.6">{desc}</div></div>',
                unsafe_allow_html=True,
            )

    with t3:
        st.markdown("### Módulos de Análise")
        modules = [
            ("📊", "#3B82F6", "Painel de Controlo",
             "Vista executiva com KPIs principais, distribuição de risco, alertas críticos e evolução mensal. "
             "Ponto de entrada ideal para a revisão diária."),
            ("🔎", "#F59E0B", "Análise de Solicitações",
             "Tabela completa de todos os solicitações com filtros por risco, prestador e período. "
             "Inclui vista em cards com detalhe de cada sinal detectado e exportação directa."),
            ("🏥", "#EF4444", "Inteligência de Prestadores",
             "Ranking de prestadores por pontuação de risco agregada. Analisa volume, valor médio, "
             "taxa de duplicados, facturação em valores redondos e padrões de fim de semana."),
            ("👤", "#A78BFA", "Análise de Beneficiários",
             "Identifica beneficiários com padrões de utilização anómalos: multi-prestadores, "
             "múltiplas solicitações no mesmo dia e gastos muito acima do grupo de referência."),
            ("💰", "#22C55E", "Custos Atípicos",
             "Compara o valor de cada solicitação com a média de pares (por procedimento, especialidade "
             "ou carteira global). Mostra scatter plot de valor vs. referência."),
            ("📋", "#F97316", "Relatório por Beneficiário",
             "Perfil de risco individual de cada beneficiário com histórico de solicitações, "
             "sinais detectados, evolução de gastos e exportação em PDF ou Excel."),
        ]
        col_a, col_b = st.columns(2)
        for i, (icon, color, name, desc) in enumerate(modules):
            with (col_a if i % 2 == 0 else col_b):
                st.markdown(
                    f'<div style="background:#1E2D3D;border:1px solid #2D3F50;border-radius:10px;'
                    f'padding:1rem 1.2rem;margin-bottom:0.7rem">'
                    f'<div style="display:flex;align-items:center;gap:0.6rem;margin-bottom:0.4rem">'
                    f'<span style="font-size:1.3rem">{icon}</span>'
                    f'<span style="font-weight:700;color:{color};font-size:0.95rem">{name}</span></div>'
                    f'<div style="font-size:0.82rem;color:#94A3B8;line-height:1.6">{desc}</div></div>',
                    unsafe_allow_html=True,
                )

    with t4:
        st.markdown("### Colunas do Ficheiro de Dados")
        st.markdown(
            '<div style="background:#2D1515;border:1px solid #EF444440;border-radius:8px;'
            'padding:0.8rem 1.1rem;margin-bottom:1rem">'
            '<strong style="color:#EF4444">Colunas obrigatórias</strong> — '
            '<span style="color:#94A3B8;font-size:0.87rem">Sem estas colunas a análise não pode ser executada.</span>'
            '</div>',
            unsafe_allow_html=True,
        )
        st.markdown("""
| Coluna | Tipo | Descrição |
|--------|------|-----------|
| `claim_id` | Texto / Número | Identificador único de cada solicitação |
| `member_id` | Texto / Número | Identificador do beneficiário |
| `provider_id` | Texto / Número | Identificador do prestador |
| `claim_date` | Data (YYYY-MM-DD) | Data de submissão do solicitação |
| `claim_amount` | Número decimal | Valor total facturado |
""")
        st.markdown(
            '<div style="background:#152D1A;border:1px solid #22C55E40;border-radius:8px;'
            'padding:0.8rem 1.1rem;margin:1rem 0">'
            '<strong style="color:#22C55E">Colunas opcionais</strong> — '
            '<span style="color:#94A3B8;font-size:0.87rem">Enriquecem a análise mas não são obrigatórias.</span>'
            '</div>',
            unsafe_allow_html=True,
        )
        st.markdown("""
| Coluna | Activa |
|--------|--------|
| `diagnosis_code` | Análise de upcoding por diagnóstico |
| `procedure_code` | Benchmark por procedimento (mais preciso) |
| `service_date` | Detecção de duplicados no mesmo dia |
| `paid_amount` | Análise de ajustes e negações |
| `provider_specialty` | Benchmark por especialidade |
| `member_age` | Segmentação demográfica |
| `member_gender` | Segmentação demográfica |
| `drug_name` / `ndc_code` | Análise de solicitações de farmácia |
| `claim_type` | Segmentação por tipo de solicitação |
""")
        st.info("A plataforma aceita nomes de colunas alternativos. Por exemplo: `npi`, `billed_amount`, `dos`, `icd10`, `cpt` — são automaticamente mapeados para os nomes canónicos.")

    st.stop()


# ── Cabeçalho da plataforma (sempre visível no Painel) ────────────────────────
if page == t("nav_overview"):
    st.markdown(
        f'<div style="display:flex;align-items:center;gap:1.1rem;'
        f'background:linear-gradient(135deg,#1E2D3D 0%,#162030 100%);'
        f'border:1px solid #2D3F50;border-left:5px solid #EF4444;'
        f'border-radius:12px;padding:1.4rem 1.8rem;margin-bottom:1.5rem">'
        f'<div style="background:#EF4444;border-radius:12px;width:56px;height:56px;'
        f'display:flex;align-items:center;justify-content:center;flex-shrink:0;'
        f'box-shadow:0 0 18px #EF444455">'
        f'<svg xmlns="http://www.w3.org/2000/svg" width="32" height="32" viewBox="0 0 24 24" fill="white">'
        f'<path d="M19 3H5a2 2 0 0 0-2 2v14a2 2 0 0 0 2 2h14a2 2 0 0 0 2-2V5a2 2 0 0 0-2-2zm-7 14h-2v-4H6v-2h4V7h2v4h4v2h-4v4z"/>'
        f'</svg></div>'
        f'<div>'
        f'<div style="font-size:2.2rem;font-weight:900;color:#F1F5F9;line-height:1.15;letter-spacing:-0.5px">'
        f'{t("platform_title")}</div>'
        f'<div style="font-size:0.95rem;color:#94A3B8;margin-top:0.5rem;letter-spacing:0.02em">'
        f'{t("platform_sub")}</div>'
        f'</div></div>',
        unsafe_allow_html=True,
    )

# ── Guardar: exige dados para páginas de análise ──────────────────────────────
if page != t("nav_data") and page != t("nav_howto") and page != t("nav_eval") and st.session_state.scored_df is None:
    if page == t("nav_overview"):
        st.markdown(
            f'<div style="background:#0D1B2A;border:1px solid #1E3A50;border-radius:10px;'
            f'padding:1.5rem 2rem;text-align:center;margin-top:1rem">'
            f'<div style="font-size:2rem;margin-bottom:0.5rem">📁</div>'
            f'<div style="color:#E2E8F0;font-size:1.05rem;font-weight:600;margin-bottom:0.4rem">'
            f'{t("no_data_title")}</div>'
            f'<div style="color:#94A3B8;font-size:0.88rem">'
            f'{t("no_data_sub")}</div>'
            f'</div>',
            unsafe_allow_html=True,
        )
    else:
        st.info(t("no_data_msg"))
    st.stop()


if st.session_state.scored_df is not None:
    df      = st.session_state.scored_df
    prov_df = st.session_state.provider_df
    mem_df  = st.session_state.member_df

    RISK_COLORS = {"High": "#EF4444", "Medium": "#F59E0B", "Low": "#22C55E"}
    RISK_PT     = {"High": t("risk_high"), "Medium": t("risk_medium"), "Low": t("risk_low")}

    # ── Painel de Controlo ─────────────────────────────────────────────────────
    if page == t("nav_overview"):
        total        = len(df)
        high         = (df["risk_level"] == "High").sum()
        medium       = (df["risk_level"] == "Medium").sum()
        low          = (df["risk_level"] == "Low").sum()
        flagged_pct  = (high + medium) / total * 100 if total > 0 else 0
        total_amt    = pd.to_numeric(df["claim_amount"], errors="coerce").sum()
        high_risk_amt = pd.to_numeric(df[df["risk_level"] == "High"]["claim_amount"], errors="coerce").sum()

        # Adjudication counts
        n_investigate = (df["adjudication"] == "Investigar Urgente").sum()   if "adjudication" in df.columns else high
        n_review      = (df["adjudication"] == "Rever Manualmente").sum()    if "adjudication" in df.columns else medium
        n_approve     = (df["adjudication"] == "Aprovar Automaticamente").sum() if "adjudication" in df.columns else low

        # ── Row 1: Risk KPIs ──────────────────────────────────────────────────
        c1, c2, c3, c4, c5 = st.columns(5)
        kpis = [
            (c1, f"{total:,}",             t("kpi_total"), "kpi-blue"),
            (c2, f"{high:,}",              t("kpi_high"),            "kpi-high"),
            (c3, f"{medium:,}",            t("kpi_medium"),           "kpi-medium"),
            (c4, f"{flagged_pct:.1f}%",    t("kpi_flagged_rate"),       "kpi-medium"),
            (c5, f"${high_risk_amt:,.0f}", t("kpi_high_value"),   "kpi-high"),
        ]
        for col, val, label, cls in kpis:
            with col:
                st.markdown(f'<div class="card"><div class="kpi-value {cls}">{val}</div>'
                            f'<div class="kpi-label">{label}</div></div>', unsafe_allow_html=True)

        # ── Row 2: Auto-Adjudication panel ───────────────────────────────────
        st.markdown(f'<div style="font-size:0.7rem;color:#475569;text-transform:uppercase;'
                    f'letter-spacing:0.08em;margin:1.2rem 0 0.5rem 0">{t("adj_panel_label")}</div>',
                    unsafe_allow_html=True)

        a1, a2, a3 = st.columns(3)
        adj_cards = [
            (a1, n_investigate, t("adj_investigate"), "#F87171", "#2D1515", "#F8717120",
             t("adj_investigate_desc")),
            (a2, n_review,      t("adj_review"),      "#FBBF24", "#2D2415", "#FBBF2420",
             t("adj_review_desc")),
            (a3, n_approve,     t("adj_approve"),    "#34D399", "#152D1A", "#34D39920",
             t("adj_approve_desc")),
        ]
        for col, count, label, color, bg, border, desc in adj_cards:
            pct = count / total * 100 if total > 0 else 0
            with col:
                st.markdown(
                    f'<div style="background:{bg};border:1px solid {border};border-radius:12px;padding:1.1rem 1.3rem">'
                    f'<div style="font-size:2rem;font-weight:800;color:{color};line-height:1">{count:,}</div>'
                    f'<div style="font-weight:700;color:{color};font-size:0.85rem;margin-top:4px">{label}</div>'
                    f'<div style="font-size:0.78rem;color:#475569;margin-top:4px">{pct:.1f}% {t("adj_pct")}</div>'
                    f'<div style="margin-top:0.6rem;height:3px;background:#1E3A50;border-radius:2px">'
                    f'<div style="width:{min(pct,100):.1f}%;height:100%;background:{color};border-radius:2px"></div></div>'
                    f'<div style="font-size:0.75rem;color:#334155;margin-top:6px">{desc}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

        st.markdown("---")

        # ── Distribuição de risco: barra + 3 cards ────────────────────────────
        high_pct   = round(high   / total * 100, 2) if total > 0 else 0
        medium_pct = round(medium / total * 100, 2) if total > 0 else 0
        low_pct    = round(low    / total * 100, 2) if total > 0 else 0
        high_mw   = "4px" if high > 0 else "0"
        medium_mw = "4px" if medium > 0 else "0"

        bar_html = (
            f'<div style="display:flex;height:28px;border-radius:8px;overflow:hidden;gap:2px;margin-bottom:1.2rem">'
            f'<div style="width:{high_pct}%;background:#F87171;min-width:{high_mw}"></div>'
            f'<div style="width:{medium_pct}%;background:#FBBF24;min-width:{medium_mw}"></div>'
            f'<div style="flex:1;background:#34D399"></div>'
            f'</div>'
        )

        def risk_card(bg, border, dot, label, count, pct, bar_color):
            return (
                f'<div style="background:{bg};border:1px solid {border};border-radius:10px;padding:1rem">'
                f'<div style="display:flex;align-items:center;gap:0.5rem;margin-bottom:0.5rem">'
                f'<div style="width:10px;height:10px;border-radius:50%;background:{dot};flex-shrink:0"></div>'
                f'<span style="color:#475569;font-size:0.72rem;text-transform:uppercase;letter-spacing:0.06em">{label}</span>'
                f'</div>'
                f'<div style="font-size:2rem;font-weight:800;color:{dot};line-height:1">{count:,}</div>'
                f'<div style="font-size:0.82rem;color:#334155;margin-top:0.3rem">{pct:.1f}% {t("adj_pct")}</div>'
                f'<div style="margin-top:0.7rem;height:4px;background:#1E3A50;border-radius:2px">'
                f'<div style="width:{max(pct,1):.1f}%;height:100%;background:{bar_color};border-radius:2px"></div>'
                f'</div>'
                f'</div>'
            )

        cards_html = (
            '<div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:0.8rem">'
            + risk_card("#2D1515", "#F8717140", "#F87171", t("high_risk_label"),  high,   high_pct,   "#F87171")
            + risk_card("#2D2415", "#FBBF2440", "#FBBF24", t("kpi_medium"), medium, medium_pct, "#FBBF24")
            + risk_card("#152D1A", "#34D39940", "#34D399", t("low_risk_label"), low,    low_pct,    "#34D399")
            + '</div>'
        )

        wrapper = (
            '<div style="background:linear-gradient(135deg,#112233,#0D1B2A);border:1px solid #1E3A50;'
            'border-radius:12px;padding:1.4rem 1.8rem;margin-bottom:1rem">'
            f'<div style="color:#475569;font-size:0.7rem;text-transform:uppercase;letter-spacing:0.08em;margin-bottom:0.8rem">'
            f'{t("risk_dist_label")} &mdash; {total:,} {t("claims_analyzed")}</div>'
            + bar_html + cards_html + '</div>'
        )
        st.markdown(wrapper, unsafe_allow_html=True)

        # ── Top 10 alerts with adjudication badge ────────────────────────────
        st.subheader(t("top10_title"))
        top10 = df[df["risk_level"] == "High"].nlargest(10, "risk_score")
        display_cols = [c for c in ["claim_id", "member_id", "provider_id", "claim_date",
                                    "claim_amount", "risk_score", "adjudication", "risk_flags"]
                        if c in top10.columns]
        for _, row in top10[display_cols].iterrows():
            score = row.get("risk_score", 0)
            amt   = row.get("claim_amount", 0)
            adj   = row.get("adjudication", t("adj_investigate"))
            # Translate adjudication label
            adj_translated = t("adj_investigate") if "Investigar" in adj or "Urgent" in adj \
                        else t("adj_review") if "Rever" in adj or "Review" in adj \
                        else t("adj_approve")
            badge_cls = "badge-investigate" if "Investigar" in adj or "Urgent" in adj \
                   else "badge-review" if "Rever" in adj or "Review" in adj \
                   else "badge-approve"
            # Translate flag text
            raw_flags = row.get("risk_flags", "")
            translated_flags = "; ".join(
                translate_flag(f.strip()) for f in raw_flags.split(";") if f.strip()
            ) if raw_flags else t("no_signals")
            st.markdown(
                f'<div class="alert-high">'
                f'<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:4px">'
                f'<strong style="color:#F1F5F9">{t("claim_label")} {row.get("claim_id","N/A")}</strong>'
                f'<span class="{badge_cls}">{adj_translated}</span></div>'
                f'<div style="display:flex;gap:1.2rem;font-size:0.82rem;flex-wrap:wrap">'
                f'<span style="color:#64748B">{t("member_label")}: <strong style="color:#CBD5E1">{row.get("member_id","N/A")}</strong></span>'
                f'<span style="color:#64748B">{t("provider_label")}: <strong style="color:#CBD5E1">{row.get("provider_id","N/A")}</strong></span>'
                f'<span style="color:#64748B">{t("value_label")}: <strong style="color:#F87171">${amt:,.2f}</strong></span>'
                f'<span style="color:#64748B">{t("score_label")}: <strong style="color:#F87171">{score:.0f}/100</strong></span>'
                f'</div>'
                f'<div style="font-size:0.78rem;color:#475569;margin-top:5px;font-style:italic">{translated_flags}</div>'
                f'</div>',
                unsafe_allow_html=True,
            )

        if "claim_date" in df.columns and df["claim_date"].notna().any():
            st.subheader(t("trend_title"))
            trend = df.groupby([df["claim_date"].dt.to_period("M"), "risk_level"], observed=True)["claim_id"].count().reset_index()
            trend["claim_date"] = trend["claim_date"].astype(str)
            fig_trend = px.bar(
                trend, x="claim_date", y="claim_id", color="risk_level",
                color_discrete_map=RISK_COLORS,
                labels={"claim_id": "Solicitações", "claim_date": "Mês", "risk_level": "Nível de Risco"},
                title="Solicitações Mensais por Nível de Risco",
            )
            fig_trend.update_layout(
                paper_bgcolor="#1E2D3D", plot_bgcolor="#1E2D3D",
                font_color="#E2E8F0", xaxis_tickangle=-45, height=320,
                margin=dict(t=40, b=60, l=10, r=10),
            )
            st.plotly_chart(fig_trend, width='stretch')


    # ── Análise de Solicitações ───────────────────────────────────────────────────
    elif page == t("nav_claims"):
        st.title(t("claims_title"))

        # ── Filtros em card ────────────────────────────────────────────────────
        st.markdown("""
        <div style="background:#1E2D3D;border:1px solid #2D3F50;border-radius:10px;
                    padding:1rem 1.4rem 0.3rem 1.4rem;margin-bottom:1rem">
            <div style="color:#94A3B8;font-size:0.7rem;text-transform:uppercase;
                        letter-spacing:0.08em;margin-bottom:0.6rem">Filtros</div>
        </div>
        """, unsafe_allow_html=True)

        fc1, fc2, fc3, fc4 = st.columns([1.2, 1, 1, 1])
        with fc1:
            risk_filter = st.multiselect(t("filter_risk"), ["High", "Medium", "Low"],
                                         default=["High", "Medium"],
                                         format_func=lambda x: RISK_PT[x])
        with fc2:
            min_score = st.slider(t("filter_min_score"), 0, 100, 40)
        with fc3:
            if "provider_id" in df.columns:
                providers = [t("filter_all")] + sorted(df["provider_id"].astype(str).unique().tolist())
                prov_filter = st.selectbox(t("filter_provider"), providers)
            else:
                prov_filter = "Todos"
        with fc4:
            if "claim_date" in df.columns and df["claim_date"].notna().any():
                min_d = df["claim_date"].min().date()
                max_d = df["claim_date"].max().date()
                date_range = st.date_input(t("filter_period"), value=(min_d, max_d),
                                           min_value=min_d, max_value=max_d)
            else:
                date_range = None

        # Apply filters
        filtered = df[df["risk_level"].isin(risk_filter) & (df["risk_score"] >= min_score)]
        if prov_filter != t("filter_all"):
            filtered = filtered[filtered["provider_id"].astype(str) == prov_filter]
        if date_range and len(date_range) == 2 and "claim_date" in df.columns:
            filtered = filtered[
                (filtered["claim_date"].dt.date >= date_range[0]) &
                (filtered["claim_date"].dt.date <= date_range[1])
            ]

        # ── Mini KPIs da filtragem ─────────────────────────────────────────────
        f_high   = (filtered["risk_level"] == "High").sum()
        f_medium = (filtered["risk_level"] == "Medium").sum()
        f_amt    = pd.to_numeric(filtered.get("claim_amount", pd.Series()), errors="coerce").sum()
        f_avg    = pd.to_numeric(filtered.get("claim_amount", pd.Series()), errors="coerce").mean()

        mk1, mk2, mk3, mk4, mk5 = st.columns(5)
        mini_kpis = [
            (mk1, f"{len(filtered):,}",  "Solicitações filtradas",   "#3B82F6"),
            (mk2, f"{f_high:,}",         t("high_risk_label"),         "#EF4444"),
            (mk3, f"{f_medium:,}",       t("kpi_medium"),        "#F59E0B"),
            (mk4, f"${f_amt:,.0f}",      "Valor Total",        "#A78BFA"),
            (mk5, f"${f_avg:,.0f}",      "Valor Médio",        "#22C55E"),
        ]
        for col, val, lbl, clr in mini_kpis:
            with col:
                st.markdown(
                    f'<div style="background:#1E2D3D;border:1px solid #2D3F50;border-top:3px solid {clr};'
                    f'border-radius:8px;padding:0.7rem 0.9rem;margin-bottom:0.5rem">'
                    f'<div style="font-size:1.4rem;font-weight:700;color:{clr}">{val}</div>'
                    f'<div style="font-size:0.72rem;color:#64748B;text-transform:uppercase;'
                    f'letter-spacing:0.05em;margin-top:2px">{lbl}</div></div>',
                    unsafe_allow_html=True,
                )

        st.markdown("---")

        # ── Visualização: cards de alto risco + tabela ─────────────────────────
        tab1, tab2 = st.tabs([t("tab_cards"), t("tab_table")])

        display_cols = [c for c in ["claim_id", "member_id", "provider_id", "claim_date",
                                     "claim_amount", "risk_score", "risk_level",
                                     "adjudication", "risk_flags"]
                        if c in filtered.columns]
        sorted_df = filtered[display_cols].sort_values("risk_score", ascending=False).head(500)

        with tab1:
            # Show top flagged solicitações as styled cards
            top_cards = sorted_df.head(30)
            if len(top_cards) == 0:
                st.info("Nenhum solicitação corresponde aos filtros seleccionados.")
            else:
                # Load existing feedback for this session
                fb_df = load_feedback(st.session_state.get("session_id", "demo"))
                fb_map = {} if fb_df.empty else dict(zip(fb_df["claim_id"], fb_df["verdict"]))

                for _, row in top_cards.iterrows():
                    score    = row.get("risk_score", 0)
                    score_lo = row.get("risk_score_low",  max(0,   score - 5))
                    score_hi = row.get("risk_score_high", min(100, score + 5))
                    fraud_prob       = row.get("fraud_probability", None)
                    supervised_label = row.get("supervised_verdict", None)
                    level   = row.get("risk_level", "Low")
                    flags   = row.get("risk_flags", "Sem sinais específicos detectados")
                    top_rf  = row.get("top_risk_factors", "")
                    amt     = row.get("claim_amount", 0)
                    cid     = str(row.get("claim_id", "N/D"))
                    clr     = "#EF4444" if level == "High" else "#F59E0B" if level == "Medium" else "#22C55E"
                    bg      = "#2D1515" if level == "High" else "#2D2415" if level == "Medium" else "#152D1A"
                    lvl_pt  = RISK_PT.get(level, level)
                    bar_w   = int(score)

                    # Existing feedback verdict (if any)
                    existing_verdict = fb_map.get(cid, "")
                    verdict_badge = ""
                    if existing_verdict == "Fraude Confirmada":
                        verdict_badge = '<span style="margin-left:0.6rem;background:#EF444422;color:#EF4444;font-size:0.68rem;font-weight:700;padding:2px 8px;border-radius:20px;border:1px solid #EF444455">✓ Fraude</span>'
                    elif existing_verdict == "Falso Positivo":
                        verdict_badge = '<span style="margin-left:0.6rem;background:#22C55E22;color:#22C55E;font-size:0.68rem;font-weight:700;padding:2px 8px;border-radius:20px;border:1px solid #22C55E55">✓ Falso +</span>'
                    elif existing_verdict == "Em Investigação":
                        verdict_badge = '<span style="margin-left:0.6rem;background:#3B82F622;color:#3B82F6;font-size:0.68rem;font-weight:700;padding:2px 8px;border-radius:20px;border:1px solid #3B82F655">🔍 Em Invest.</span>'

                    # CI display
                    ci_str = f"IC: {score_lo:.0f}–{score_hi:.0f}"

                    # Supervised model badge
                    sup_badge = ""
                    if fraud_prob is not None:
                        try:
                            fp_val = float(fraud_prob)
                            if supervised_label == "Provável Fraude":
                                sup_clr = "#EF4444"
                            elif supervised_label == "Incerto":
                                sup_clr = "#A78BFA"
                            else:
                                sup_clr = "#22C55E"
                            sup_badge = (
                                f'<span style="margin-left:0.6rem;background:{sup_clr}22;color:{sup_clr};'
                                f'font-size:0.68rem;font-weight:700;padding:2px 8px;border-radius:20px;'
                                f'border:1px solid {sup_clr}55">🤖 {supervised_label} ({fp_val:.0f}%)</span>'
                            )
                        except (TypeError, ValueError):
                            pass

                    st.markdown(
                        f'<div style="background:{bg};border:1px solid {clr}33;border-left:4px solid {clr};'
                        f'border-radius:10px;padding:0.9rem 1.2rem;margin-bottom:0.5rem">'
                        f'<div style="display:flex;justify-content:space-between;align-items:flex-start">'
                        f'<div>'
                        f'<span style="font-size:1rem;font-weight:700;color:#F1F5F9">Solicitação {cid}</span>'
                        f'<span style="margin-left:0.8rem;background:{clr}22;color:{clr};font-size:0.72rem;'
                        f'font-weight:700;padding:2px 10px;border-radius:20px;border:1px solid {clr}55">{lvl_pt}</span>'
                        f'{sup_badge}'
                        f'{verdict_badge}'
                        f'</div>'
                        f'<div style="text-align:right">'
                        f'<div style="font-size:1.3rem;font-weight:800;color:{clr}">{score:.0f}<span style="font-size:0.7rem;color:#64748B">/100</span></div>'
                        f'<div style="font-size:0.68rem;color:#475569">{ci_str}</div>'
                        f'</div>'
                        f'</div>'
                        f'<div style="margin-top:0.4rem;height:3px;background:#2D3F50;border-radius:2px">'
                        f'<div style="width:{bar_w}%;height:100%;background:{clr};border-radius:2px"></div>'
                        f'</div>'
                        f'<div style="display:flex;gap:1.5rem;margin-top:0.5rem;flex-wrap:wrap">'
                        f'<span style="color:#64748B;font-size:0.8rem">Benef.: <strong style="color:#CBD5E1">{row.get("member_id","N/D")}</strong></span>'
                        f'<span style="color:#64748B;font-size:0.8rem">Prestador: <strong style="color:#CBD5E1">{row.get("provider_id","N/D")}</strong></span>'
                        f'<span style="color:#64748B;font-size:0.8rem">Valor: <strong style="color:#CBD5E1">${amt:,.2f}</strong></span>'
                        f'<span style="color:#64748B;font-size:0.8rem">Data: <strong style="color:#CBD5E1">{str(row.get("claim_date",""))[:10]}</strong></span>'
                        f'</div>'
                        f'{"<div style=margin-top:0.45rem;background:#1a2535;border-radius:6px;padding:0.4rem 0.7rem;font-size:0.75rem;color:#94A3B8><strong style=color:#F59E0B>Top fatores de risco:</strong> " + top_rf + "</div>" if top_rf else ""}'
                        f'<div style="margin-top:0.35rem;font-size:0.78rem;color:#94A3B8;font-style:italic">{flags}</div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )

                    # Feedback buttons — compact row under each card
                    fb_col1, fb_col2, fb_col3, fb_col4 = st.columns([1.6, 1.4, 1.4, 5])
                    _sid   = st.session_state.get("session_id", "demo")
                    _conf  = st.session_state.get(f"conf_{cid}", 3)
                    with fb_col1:
                        if st.button("🚨 Fraude Confirmada", key=f"fb_fraud_{cid}", use_container_width=True):
                            save_feedback(cid, _sid, "Fraude Confirmada", confidence_level=_conf)
                            st.rerun()
                    with fb_col2:
                        if st.button("✅ Falso Positivo", key=f"fb_fp_{cid}", use_container_width=True):
                            save_feedback(cid, _sid, "Falso Positivo", confidence_level=_conf)
                            st.rerun()
                    with fb_col3:
                        if st.button("🔍 Em Investigação", key=f"fb_inv_{cid}", use_container_width=True):
                            save_feedback(cid, _sid, "Em Investigação", confidence_level=_conf)
                            st.rerun()
                    with fb_col4:
                        st.select_slider(
                            "Confiança na decisão",
                            options=[1, 2, 3, 4, 5],
                            format_func=lambda v: {1:"⭐ Baixa", 2:"⭐⭐", 3:"⭐⭐⭐ Média", 4:"⭐⭐⭐⭐", 5:"⭐⭐⭐⭐⭐ Alta"}[v],
                            key=f"conf_{cid}",
                            label_visibility="collapsed",
                        )

        with tab2:
            def style_risk(val):
                if val == "High":   return "background-color:#2D1515;color:#EF4444;font-weight:bold"
                elif val == "Medium": return "background-color:#2D2415;color:#F59E0B;font-weight:bold"
                return "background-color:#152D1A;color:#22C55E;font-weight:bold"

            if "risk_level" in sorted_df.columns:
                st.dataframe(
                    sorted_df.style.map(style_risk, subset=["risk_level"]),
                    width='stretch', height=520,
                )
            else:
                st.dataframe(sorted_df, width='stretch', height=520)

        # ── AI Status line (simple, user-facing) ─────────────────────────────
        sm_metrics  = st.session_state.get("supervised_metrics")
        fb_for_auto = load_feedback(st.session_state.get("session_id", "demo"))

        # ── Auto-train: fires when enough new labels accumulate ──────────────
        _n_fraud_auto = 0 if fb_for_auto.empty else (fb_for_auto["verdict"] == "Fraude Confirmada").sum()
        _n_legit_auto = 0 if fb_for_auto.empty else (fb_for_auto["verdict"] == "Falso Positivo").sum()
        _prev_labeled = (sm_metrics or {}).get("n_labeled", 0)
        _new_total    = int(_n_fraud_auto) + int(_n_legit_auto)
        _did_retrain  = False
        if _n_fraud_auto >= 5 and _n_legit_auto >= 5 and _new_total > _prev_labeled:
            _auto_model, _auto_metrics = supervised_model.train(
                st.session_state.scored_df, fb_for_auto
            )
            if _auto_metrics and _auto_metrics.get("status") == "trained":
                st.session_state.scored_df = supervised_model.predict(
                    st.session_state.scored_df, _auto_model
                )
                st.session_state.supervised_metrics = _auto_metrics
                sm_metrics   = _auto_metrics
                _did_retrain = True

        # ── Detect external model update (Jupyter / model_training.py) ───────
        _pkl = supervised_model.MODEL_PATH
        _current_mtime = _pkl.stat().st_mtime if _pkl.exists() else 0
        if _current_mtime > st.session_state.get("model_pkl_mtime", 0):
            st.session_state.model_pkl_mtime = _current_mtime
            _ext_metrics = supervised_model.load_metrics()
            if _ext_metrics and _ext_metrics.get("status") == "trained":
                st.session_state.scored_df = supervised_model.predict(
                    st.session_state.scored_df
                )
                st.session_state.supervised_metrics = _ext_metrics
                sm_metrics   = _ext_metrics
                _did_retrain = True

        # ── Re-score notification + page refresh ──────────────────────────────
        if _did_retrain:
            _n_lab_new = (sm_metrics or {}).get("n_labeled", 0)
            st.toast(
                f"🤖 Modelo actualizado — scores re-calculados com {_n_lab_new} avaliações",
                icon="✅"
            )
            st.rerun()

        # One quiet status line — all the user needs to see
        st.markdown("---")
        if sm_metrics and sm_metrics.get("status") == "trained":
            n_lab = sm_metrics.get("n_labeled", 0)
            st.markdown(
                f'<div style="background:#1E2D3D;border-radius:6px;padding:0.55rem 1rem;'
                f'font-size:0.82rem;color:#64748B">'
                f'🤖 <strong style="color:#22C55E">IA activa</strong> — '
                f'aprendeu com <strong style="color:#E2E8F0">{n_lab}</strong> avaliações de investigadores. '
                f'As probabilidades de fraude acima são actualizadas automaticamente.</div>',
                unsafe_allow_html=True
            )
        elif _n_fraud_auto > 0 or _n_legit_auto > 0:
            _need = max(0, 5 - int(_n_fraud_auto)) + max(0, 5 - int(_n_legit_auto))
            st.markdown(
                f'<div style="background:#1E2D3D;border-radius:6px;padding:0.55rem 1rem;'
                f'font-size:0.82rem;color:#64748B">'
                f'🤖 <strong style="color:#F59E0B">IA em aprendizagem</strong> — '
                f'{int(_n_fraud_auto) + int(_n_legit_auto)} avaliações registadas. '
                f'Mais {_need} para activar as probabilidades de fraude automáticas.</div>',
                unsafe_allow_html=True
            )
        else:
            st.markdown(
                '<div style="background:#1E2D3D;border-radius:6px;padding:0.55rem 1rem;'
                'font-size:0.82rem;color:#64748B">'
                '🤖 <strong style="color:#94A3B8">IA inactiva</strong> — '
                'marque solicitações acima como Fraude ou Falso Positivo para activar.</div>',
                unsafe_allow_html=True
            )

        # ── Feedback summary — counts only, no technical detail ───────────────
        fb_all = load_feedback(st.session_state.get("session_id", "demo"))
        if not fb_all.empty:
            v_counts  = fb_all["verdict"].value_counts()
            confirmed = v_counts.get("Fraude Confirmada", 0)
            fp_count  = v_counts.get("Falso Positivo", 0)
            inv_count = v_counts.get("Em Investigação", 0)
            fc1, fc2, fc3 = st.columns(3)
            fc1.metric("🚨 Fraude Confirmada", confirmed)
            fc2.metric("✅ Falso Positivo",     fp_count)
            fc3.metric("🔍 Em Investigação",    inv_count)

            with st.expander("Ver todos os registos de feedback"):
                st.dataframe(
                    fb_all[["claim_id", "verdict", "investigator", "notes", "created_at"]],
                    width="stretch"
                )

        # ── Exportar ───────────────────────────────────────────────────────────
        st.markdown("---")
        st.subheader(t("export_title"))
        ec1, ec2, ec3 = st.columns(3)
        with ec1:
            xlsx = exporter.to_excel(df, prov_df, mem_df, df)
            st.download_button(t("btn_excel"), xlsx, "relatorio_claims.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
        with ec2:
            csv_bytes = exporter.to_csv(df)
            st.download_button(t("btn_csv"), csv_bytes, "claims_sinalizados.csv",
                               "text/csv", use_container_width=True)
        with ec3:
            pdf_bytes = exporter.to_pdf(df, prov_df, mem_df)
            st.download_button(t("btn_pdf"), pdf_bytes, "relatorio_investigacao.pdf",
                               "application/pdf", use_container_width=True)


    # ── Inteligência de Prestadores ────────────────────────────────────────────
    elif page == t("nav_providers"):
        st.title(t("prov_title"))

        if prov_df is None or len(prov_df) == 0:
            st.warning("Dados de prestadores não disponíveis.")
            st.stop()

        top_n = st.slider(t("prov_slider"), 5, 50, 20)
        top_prov = prov_df.nlargest(top_n, "provider_risk_score")

        top_prov_sorted = top_prov.sort_values("provider_risk_score", ascending=True)
        prov_colors = [
            "#EF4444" if v >= 70 else "#F59E0B" if v >= 40 else "#22C55E"
            for v in top_prov_sorted["provider_risk_score"]
        ]
        fig_bar = go.Figure(go.Bar(
            x=top_prov_sorted["provider_risk_score"],
            y=top_prov_sorted["provider_id"].astype(str),
            orientation="h",
            marker=dict(color=prov_colors, line=dict(color="#0F1923", width=1)),
            text=[f"{v:.0f}" for v in top_prov_sorted["provider_risk_score"]],
            textposition="outside",
            textfont=dict(color="#E2E8F0", size=11),
            hovertemplate="<b>%{y}</b><br>Pontuação: %{x:.1f}<extra></extra>",
        ))
        fig_bar.update_layout(
            title=dict(
                text=f"Top {top_n} Prestadores por Pontuação de Risco",
                font=dict(size=14, color="#94A3B8"), x=0.5, xanchor="center",
            ),
            paper_bgcolor="#1E2D3D", plot_bgcolor="#1E2D3D",
            font_color="#E2E8F0",
            xaxis=dict(
                range=[0, 115], showgrid=True, gridcolor="#243447",
                zeroline=False, tickfont=dict(color="#64748B"),
                title=dict(text="Pontuação de Risco", font=dict(size=11, color="#64748B")),
            ),
            yaxis=dict(showgrid=False, tickfont=dict(color="#CBD5E1", size=11)),
            margin=dict(t=50, b=20, l=120, r=60),
            height=max(320, top_n * 28),
            showlegend=False,
        )
        st.plotly_chart(fig_bar, width='stretch')

        st.subheader(t("prov_detail"))
        sel_provider = st.selectbox(t("prov_select"), prov_df["provider_id"].astype(str).tolist())
        prov_row = prov_df[prov_df["provider_id"].astype(str) == sel_provider]

        if len(prov_row) > 0:
            r = prov_row.iloc[0]
            p1, p2, p3, p4 = st.columns(4)
            p1.metric(t("prov_risk_score"),  f"{r.get('provider_risk_score', 0):.1f}")
            p2.metric(t("kpi_total"),  f"{int(r.get('claim_count', 0)):,}")
            p3.metric(t("prov_avg_value"),         f"${r.get('avg_amount', 0):,.2f}")
            p4.metric(t("prov_dup_rate"),  f"{r.get('dup_rate', 0)*100:.1f}%")

            flags = r.get("provider_flags", "")
            if flags:
                for f in flags.split(";"):
                    if f.strip():
                        st.markdown(f'<div class="alert-high">{f.strip()}</div>', unsafe_allow_html=True)

            provider_claims = df[df["provider_id"].astype(str) == sel_provider]
            if len(provider_claims) > 0:
                st.caption(f"{len(provider_claims):,} solicitações deste prestador")
                show_cols = [c for c in ["claim_id", "member_id", "claim_date",
                                          "claim_amount", "risk_score", "risk_level"]
                             if c in provider_claims.columns]
                st.dataframe(provider_claims[show_cols].sort_values("risk_score", ascending=False).head(100),
                             width='stretch')

            # ── Provider Monthly Trend Lines ──────────────────────────────────
            trends_df = st.session_state.get("trends_df")
            if trends_df is not None and not trends_df.empty:
                prov_trends = trends_df[trends_df["provider_id"].astype(str) == sel_provider]
                if len(prov_trends) >= 2:
                    st.markdown("#### 📈 Tendência Mensal — Prestador Seleccionado")
                    tc1, tc2 = st.columns(2)

                    with tc1:
                        fig_t1 = go.Figure()
                        fig_t1.add_trace(go.Scatter(
                            x=prov_trends["month"], y=prov_trends["claim_count"],
                            mode="lines+markers",
                            line=dict(color="#3B82F6", width=2),
                            marker=dict(size=7, color="#3B82F6"),
                            name="Nº de Solicitações",
                            hovertemplate="<b>%{x}</b><br>Claims: %{y}<extra></extra>",
                        ))
                        fig_t1.update_layout(
                            title=dict(text="Volume de Solicitações/Mês", font=dict(size=12, color="#94A3B8"), x=0.5, xanchor="center"),
                            paper_bgcolor="#1E2D3D", plot_bgcolor="#1E2D3D",
                            font_color="#E2E8F0", height=260,
                            xaxis=dict(showgrid=False, tickfont=dict(color="#64748B", size=10)),
                            yaxis=dict(showgrid=True, gridcolor="#243447", tickfont=dict(color="#64748B")),
                            margin=dict(t=40, b=30, l=50, r=20),
                        )
                        st.plotly_chart(fig_t1, use_container_width=True)

                    with tc2:
                        fig_t2 = go.Figure()
                        fig_t2.add_trace(go.Scatter(
                            x=prov_trends["month"], y=prov_trends["avg_amount"],
                            mode="lines+markers",
                            line=dict(color="#F59E0B", width=2),
                            marker=dict(size=7, color="#F59E0B"),
                            name="Valor Médio",
                            hovertemplate="<b>%{x}</b><br>Valor Médio: $%{y:,.2f}<extra></extra>",
                        ))
                        fig_t2.add_trace(go.Bar(
                            x=prov_trends["month"], y=prov_trends["dup_count"],
                            name="Duplicados",
                            marker=dict(color="rgba(239,68,68,0.4)"),
                            yaxis="y2",
                            hovertemplate="<b>%{x}</b><br>Duplicados: %{y}<extra></extra>",
                        ))
                        fig_t2.update_layout(
                            title=dict(text="Valor Médio/Mês + Duplicados", font=dict(size=12, color="#94A3B8"), x=0.5, xanchor="center"),
                            paper_bgcolor="#1E2D3D", plot_bgcolor="#1E2D3D",
                            font_color="#E2E8F0", height=260,
                            xaxis=dict(showgrid=False, tickfont=dict(color="#64748B", size=10)),
                            yaxis=dict(showgrid=True, gridcolor="#243447", tickfont=dict(color="#64748B"), title="Valor Médio ($)"),
                            yaxis2=dict(overlaying="y", side="right", tickfont=dict(color="#EF4444", size=9), title="Duplicados"),
                            legend=dict(orientation="h", y=-0.2, font=dict(size=10)),
                            margin=dict(t=40, b=30, l=60, r=50),
                        )
                        st.plotly_chart(fig_t2, use_container_width=True)
                elif len(prov_trends) == 1:
                    st.info("Apenas 1 mês de dados disponível para este prestador — tendência disponível com 2+ meses.")

        if "avg_amount" in prov_df.columns and "claim_count" in prov_df.columns:
            fig_scatter = px.scatter(
                prov_df,
                x="claim_count", y="avg_amount",
                size="provider_risk_score", color="provider_risk_score",
                hover_data=["provider_id"],
                color_continuous_scale=[[0, "#22C55E"], [0.5, "#F59E0B"], [1, "#EF4444"]],
                title="Risco do Prestador: Volume de Solicitações vs. Valor Médio",
                labels={"claim_count": "Número de Solicitações", "avg_amount": "Valor Médio ($)"},
            )
            fig_scatter.update_layout(
                paper_bgcolor="#1E2D3D", plot_bgcolor="#1E2D3D",
                font_color="#E2E8F0", height=400,
                coloraxis_colorbar_title="Risco",
            )
            st.plotly_chart(fig_scatter, width='stretch')


    # ── Análise de Beneficiários ───────────────────────────────────────────────
    elif page == t("nav_members"):
        st.title(t("mem_title"))

        if mem_df is None or len(mem_df) == 0:
            st.warning("Dados de beneficiários não disponíveis.")
            st.stop()

        top_n_mem = st.slider(t("mem_slider"), 5, 50, 20)
        top_mem = mem_df.nlargest(top_n_mem, "member_risk_score")

        top_mem_sorted = top_mem.sort_values("member_risk_score", ascending=True)
        bar_colors = [
            "#EF4444" if v >= 70 else "#F59E0B" if v >= 40 else "#22C55E"
            for v in top_mem_sorted["member_risk_score"]
        ]
        fig_mem = go.Figure(go.Bar(
            x=top_mem_sorted["member_risk_score"],
            y=top_mem_sorted["member_id"].astype(str),
            orientation="h",
            marker=dict(color=bar_colors, line=dict(color="#0F1923", width=1)),
            text=[f"{v:.0f}" for v in top_mem_sorted["member_risk_score"]],
            textposition="outside",
            textfont=dict(color="#E2E8F0", size=11),
            hovertemplate="<b>%{y}</b><br>Pontuação: %{x:.1f}<extra></extra>",
        ))
        fig_mem.update_layout(
            title=dict(
                text=f"Top {top_n_mem} Beneficiários por Pontuação de Risco",
                font=dict(size=14, color="#94A3B8"), x=0.5, xanchor="center",
            ),
            paper_bgcolor="#1E2D3D", plot_bgcolor="#1E2D3D",
            font_color="#E2E8F0",
            xaxis=dict(
                range=[0, 115], showgrid=True, gridcolor="#243447",
                zeroline=False, tickfont=dict(color="#64748B"),
                title=dict(text="Pontuação de Risco", font=dict(size=11, color="#64748B")),
            ),
            yaxis=dict(showgrid=False, tickfont=dict(color="#CBD5E1", size=11)),
            margin=dict(t=50, b=20, l=120, r=60),
            height=max(320, top_n_mem * 28),
            showlegend=False,
        )
        st.plotly_chart(fig_mem, width='stretch')

        if "distinct_providers" in mem_df.columns and "total_spend" in mem_df.columns:
            fig_shop = px.scatter(
                mem_df,
                x="distinct_providers", y="total_spend",
                color="member_risk_score",
                size="claim_count",
                hover_data=["member_id"],
                color_continuous_scale=[[0, "#22C55E"], [0.5, "#F59E0B"], [1, "#EF4444"]],
                title="Padrão de Utilização: Prestadores Distintos vs. Gasto Total",
                labels={"distinct_providers": "Prestadores Distintos", "total_spend": "Gasto Total ($)"},
            )
            fig_shop.update_layout(
                paper_bgcolor="#1E2D3D", plot_bgcolor="#1E2D3D",
                font_color="#E2E8F0", height=400,
            )
            st.plotly_chart(fig_shop, width='stretch')

        st.subheader(t("mem_table_title"))
        m_cols = [c for c in ["member_id", "member_risk_score", "claim_count", "total_spend",
                               "distinct_providers", "member_flags"] if c in mem_df.columns]
        st.dataframe(
            mem_df[m_cols].sort_values("member_risk_score", ascending=False).head(200),
            width='stretch',
        )


    # ── Custos Atípicos ────────────────────────────────────────────────────────
    elif page == t("nav_costs"):
        st.title(t("cost_title"))

        if "cost_outlier_score" not in df.columns:
            st.warning("Pontuações de custos atípicos não disponíveis.")
            st.stop()

        profile = st.session_state.profile
        benchmark_label = (
            "código de procedimento" if profile.has_procedure_code else
            "especialidade do prestador" if profile.has_provider_specialty else
            "carteira global"
        )
        st.info(f"Grupo de referência (benchmark): **{benchmark_label}**")

        scatter_df = df.copy()
        scatter_df["_amount"] = pd.to_numeric(scatter_df["claim_amount"], errors="coerce")

        if "peer_mean" in scatter_df.columns:
            fig_scatter = px.scatter(
                scatter_df,
                x="peer_mean", y="_amount",
                color="risk_level",
                color_discrete_map=RISK_COLORS,
                hover_data=[c for c in ["claim_id", "provider_id", "risk_score"] if c in scatter_df.columns],
                title="Valor do Solicitação vs. Referência de Pares",
                labels={"peer_mean": "Média de Pares ($)", "_amount": "Valor Facturado ($)",
                        "risk_level": "Nível de Risco"},
                opacity=0.7,
            )
            max_val = max(scatter_df["_amount"].max(), scatter_df["peer_mean"].max())
            fig_scatter.add_shape(type="line", x0=0, y0=0, x1=max_val, y1=max_val,
                                  line=dict(color="#3B82F6", dash="dash", width=1))
            fig_scatter.update_layout(
                paper_bgcolor="#1E2D3D", plot_bgcolor="#1E2D3D",
                font_color="#E2E8F0", height=450,
            )
            st.plotly_chart(fig_scatter, width='stretch')

        fig_dist = px.histogram(
            scatter_df, x="_amount", nbins=60,
            color="risk_level",
            color_discrete_map=RISK_COLORS,
            title="Distribuição dos Valores dos Solicitações por Nível de Risco",
            labels={"_amount": "Valor do Solicitação ($)", "risk_level": "Nível de Risco"},
            barmode="overlay", opacity=0.8,
        )
        fig_dist.update_layout(
            paper_bgcolor="#1E2D3D", plot_bgcolor="#1E2D3D",
            font_color="#E2E8F0", height=350,
        )
        st.plotly_chart(fig_dist, width='stretch')

        st.subheader(t("cost_table_title"))
        outlier_cols = [c for c in ["claim_id", "member_id", "provider_id", "claim_amount",
                                     "peer_mean", "cost_outlier_score", "risk_level", "risk_flags"]
                        if c in df.columns]
        top_outliers = df.nlargest(100, "cost_outlier_score")[outlier_cols]
        st.dataframe(top_outliers, width='stretch')


    # ── Relatório por Beneficiário ─────────────────────────────────────────────
    elif page == t("nav_report"):
        st.title(t("report_title"))
        st.caption("Seleccione um beneficiário para gerar o seu perfil de risco completo e exportar em PDF ou Excel.")

        if mem_df is None or len(mem_df) == 0:
            st.warning("Dados de beneficiários não disponíveis.")
            st.stop()

        # Selector
        RISK_PT = {"High": "Alto", "Medium": "Médio", "Low": "Baixo"}
        members_sorted = mem_df.sort_values("member_risk_score", ascending=False)["member_id"].astype(str).tolist()
        sel_member = st.selectbox("Seleccionar Beneficiário", members_sorted,
                                  format_func=lambda m: f"{m}  —  Risco: {mem_df[mem_df['member_id'].astype(str)==m]['member_risk_score'].values[0]:.0f}/100")

        mem_row = mem_df[mem_df["member_id"].astype(str) == sel_member]
        mem_claims = df[df["member_id"].astype(str) == sel_member].sort_values("risk_score", ascending=False)

        if len(mem_row) == 0:
            st.warning("Beneficiário não encontrado.")
            st.stop()

        r = mem_row.iloc[0]
        risk_score = r.get("member_risk_score", 0)
        risk_level = t("risk_high") if risk_score >= 70 else t("risk_medium") if risk_score >= 40 else t("risk_low")
        risk_color = "#EF4444" if risk_score >= 70 else "#F59E0B" if risk_score >= 40 else "#22C55E"
        risk_bg    = "#2D1515" if risk_score >= 70 else "#2D2415" if risk_score >= 40 else "#152D1A"

        # ── Cabeçalho do relatório ─────────────────────────────────────────────
        st.markdown(f"""
        <div style="background:#1E2D3D;border:1px solid #2D3F50;border-left:5px solid {risk_color};
                    border-radius:12px;padding:1.4rem 1.8rem;margin-bottom:1.2rem;display:flex;
                    align-items:center;gap:1.5rem">
            <div style="background:{risk_bg};border:2px solid {risk_color};border-radius:50%;
                        width:64px;height:64px;display:flex;align-items:center;justify-content:center;
                        font-size:1.6rem;font-weight:800;color:{risk_color};flex-shrink:0">{risk_score:.0f}</div>
            <div>
                <div style="font-size:1.3rem;font-weight:700;color:#F1F5F9">{t("member_label")}: {sel_member}</div>
                <div style="font-size:0.9rem;color:{risk_color};font-weight:600;margin-top:2px">
                    {t("report_risk_label")} {risk_level}</div>
                <div style="font-size:0.82rem;color:#64748B;margin-top:2px">
                    {int(r.get('claim_count',0)):,} {t("sidebar_claims")} &bull;
                    {t("report_total_spend")}: ${r.get('total_spend',0):,.2f} &bull;
                    {int(r.get('distinct_providers',0))} {t("report_providers")}</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # ── KPIs do beneficiário ───────────────────────────────────────────────
        k1, k2, k3, k4 = st.columns(4)
        k1.metric(t("prov_risk_score"),      f"{risk_score:.1f} / 100")
        k2.metric(t("kpi_total"),         f"{int(r.get('claim_count', 0)):,}")
        k3.metric(t("report_total_spend"),             f"${r.get('total_spend', 0):,.2f}")
        k4.metric(t("report_providers"),   f"{int(r.get('distinct_providers', 0))}")

        st.markdown("---")

        # ── Sinais de risco ───────────────────────────────────────────────────
        flags_raw = r.get("member_flags", "")
        flags_list = [f.strip() for f in flags_raw.split(";") if f.strip()]

        if flags_list:
            st.subheader(t("report_flags"))
            for flag in flags_list:
                alert_cls = "alert-high" if risk_score >= 70 else "alert-medium"
                st.markdown(f'<div class="{alert_cls}">&#9888; {flag}</div>', unsafe_allow_html=True)
        else:
            st.success("Nenhum sinal de risco específico detectado para este beneficiário.")

        st.markdown("---")

        # ── Solicitações do beneficiário ────────────────────────────────────────────
        _claims_lbl = t("report_claims_title")
        _recs_lbl   = t("data_records")
        st.subheader(f"{_claims_lbl} ({len(mem_claims):,} {_recs_lbl})")

        show_cols = [c for c in ["claim_id", "provider_id", "claim_date", "claim_amount",
                                  "risk_score", "risk_level", "risk_flags"] if c in mem_claims.columns]

        def style_risk(val):
            if val == "High":   return "background-color:#2D1515;color:#EF4444;font-weight:bold"
            elif val == "Medium": return "background-color:#2D2415;color:#F59E0B;font-weight:bold"
            return "background-color:#152D1A;color:#22C55E;font-weight:bold"

        if "risk_level" in mem_claims.columns:
            st.dataframe(
                mem_claims[show_cols].style.map(style_risk, subset=["risk_level"]),
                width='stretch', height=350,
            )
        else:
            st.dataframe(mem_claims[show_cols], width='stretch', height=350)

        # ── Gráfico: evolução de gastos ───────────────────────────────────────
        if "claim_date" in mem_claims.columns and mem_claims["claim_date"].notna().any():
            st.subheader(t("report_trend"))
            timeline = mem_claims.groupby(
                mem_claims["claim_date"].dt.to_period("M"), observed=True
            )["claim_amount"].sum().reset_index()
            timeline["claim_date"] = timeline["claim_date"].astype(str)
            fig_tl = px.bar(
                timeline, x="claim_date", y="claim_amount",
                labels={"claim_date": "Mês", "claim_amount": "Valor ($)"},
                title="Gasto Mensal do Beneficiário",
                color_discrete_sequence=["#3B82F6"],
            )
            fig_tl.update_layout(
                paper_bgcolor="#1E2D3D", plot_bgcolor="#1E2D3D",
                font_color="#E2E8F0", height=280,
                xaxis_tickangle=-45,
                margin=dict(t=40, b=50, l=10, r=10),
            )
            st.plotly_chart(fig_tl, width='stretch')

        st.markdown("---")

        # ── Exportar relatório individual ─────────────────────────────────────
        st.subheader(t("report_export"))
        ex1, ex2 = st.columns(2)

        with ex1:
            # PDF individual do beneficiário
            from modules.exporter import to_pdf
            member_scored = df[df["member_id"].astype(str) == sel_member]
            pdf_bytes = to_pdf(member_scored, prov_df, mem_df)
            st.download_button(
                t("report_pdf_btn"),
                pdf_bytes,
                f"relatorio_{sel_member}.pdf",
                "application/pdf",
                use_container_width=True,
            )

        with ex2:
            # Excel individual
            from io import BytesIO
            from openpyxl import Workbook as OWorkbook
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter

            def member_excel(member_id, claims_df, mem_info):
                wb = OWorkbook()
                ws = wb.active
                ws.title = "Perfil de Risco"
                navy = PatternFill("solid", fgColor="FF0F1923")
                white_font = Font(color="FFFFFFFF", bold=True)

                # Summary block
                ws.append(["RELATÓRIO DE RISCO — BENEFICIÁRIO", member_id])
                ws.append(["Pontuação de Risco", f"{mem_info.get('member_risk_score', 0):.1f} / 100"])
                ws.append(["Nível de Risco", risk_level])
                ws.append([t("kpi_total"), int(mem_info.get("claim_count", 0))])
                ws.append(["Gasto Total", f"${mem_info.get('total_spend', 0):,.2f}"])
                ws.append(["Prestadores Distintos", int(mem_info.get("distinct_providers", 0))])
                ws.append(["Sinais de Risco", flags_raw or "Nenhum"])
                ws.append([])

                # Solicitações table
                cols = [c for c in ["claim_id", "provider_id", "claim_date", "claim_amount",
                                    "risk_score", "risk_level", "risk_flags"] if c in claims_df.columns]
                ws.append(cols)
                for cell in ws[ws.max_row]:
                    cell.fill = navy
                    cell.font = white_font

                for row_data in claims_df[cols].itertuples(index=False):
                    ws.append(list(row_data))

                for i in range(1, len(cols) + 1):
                    ws.column_dimensions[get_column_letter(i)].width = 20

                buf = BytesIO()
                wb.save(buf)
                return buf.getvalue()

            xlsx_bytes = member_excel(sel_member, mem_claims, r)
            st.download_button(
                t("report_xlsx_btn"),
                xlsx_bytes,
                f"relatorio_{sel_member}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

# ══════════════════════════════════════════════════════════════════════════════
# PÁGINA: Avaliação da Plataforma (não requer dados carregados)
# ══════════════════════════════════════════════════════════════════════════════
if page == t("nav_eval"):
    st.title(f"⭐ {t('eval_title')}")
    st.caption(t("eval_caption"))

    tab_form, tab_results = st.tabs([t("eval_tab_form"), t("eval_tab_results")])

    # ── Tab 1: Formulário ─────────────────────────────────────────────────────
    with tab_form:
        st.markdown(
            f"<div style='background:#1E2D3D;border:1px solid #2D3F50;border-left:4px solid #F59E0B;"
            f"border-radius:10px;padding:1rem 1.5rem;margin-bottom:1.5rem'>"
            f"<div style='color:#F59E0B;font-weight:700;margin-bottom:4px'>{t('eval_info_title')}</div>"
            f"<div style='color:#94A3B8;font-size:0.87rem'>{t('eval_info_body')}</div>"
            f"</div>",
            unsafe_allow_html=True,
        )

        _role_opts    = t("eval_role_opts")
        _role_none    = _role_opts[0]
        _time_opts    = t("eval_time_opts")
        _rec_opts     = t("eval_rec_opts")
        _useful_lbls  = t("eval_useful_labels")
        _acc_lbls     = t("eval_acc_labels")

        with st.form("eval_form", clear_on_submit=True):
            st.subheader(t("eval_section_id"))
            ev_c1, ev_c2 = st.columns(2)
            with ev_c1:
                ev_name = st.text_input(t("eval_name"), placeholder=t("eval_name_ph"))
            with ev_c2:
                ev_role = st.selectbox(t("eval_role"), _role_opts)

            st.markdown("---")
            st.subheader(t("eval_section_eff"))

            ev_useful = st.slider(
                t("eval_useful_q"),
                min_value=1, max_value=5, value=3,
                format="%d ⭐",
                help=t("eval_useful_help"),
            )
            st.caption(f"{t('eval_selected')} **{_useful_lbls[ev_useful]}**")

            ev_time = st.radio(t("eval_time_q"), _time_opts, horizontal=False)

            ev_accuracy = st.slider(
                t("eval_acc_q"),
                min_value=1, max_value=5, value=3,
                format="%d ⭐",
                help=t("eval_acc_help"),
            )
            st.caption(f"{t('eval_selected')} **{_acc_lbls[ev_accuracy]}**")

            st.markdown("---")
            st.subheader(t("eval_section_rec"))

            ev_recommend = st.radio(t("eval_rec_q"), _rec_opts, horizontal=True)

            ev_features = st.text_area(
                t("eval_feat_q"),
                placeholder=t("eval_feat_ph"),
                height=100,
            )

            ev_comments = st.text_area(
                t("eval_comments_q"),
                placeholder=t("eval_comments_ph"),
                height=120,
            )

            st.markdown("")
            submitted = st.form_submit_button(
                t("eval_submit"),
                use_container_width=True,
                type="primary",
            )

        if submitted:
            role_val = ev_role if ev_role != _role_none else ""
            save_evaluation(
                respondent_name   = ev_name.strip(),
                respondent_role   = role_val,
                usefulness_rating = ev_useful,
                time_savings      = ev_time,
                accuracy_rating   = ev_accuracy,
                recommendation    = ev_recommend,
                feature_requests  = ev_features.strip(),
                general_comments  = ev_comments.strip(),
                session_id        = st.session_state.get("session_id"),
            )
            st.success(t("eval_success"))
            st.balloons()

    # ── Tab 2: Resultados ─────────────────────────────────────────────────────
    with tab_results:
        eval_df = load_evaluations()

        if eval_df.empty:
            st.info(t("eval_empty"))
        else:
            n_resp = len(eval_df)

            avg_useful   = eval_df["usefulness_rating"].dropna().mean()
            avg_accuracy = eval_df["accuracy_rating"].dropna().mean()

            rk1, rk2, rk3 = st.columns(3)
            rk1.metric(t("eval_kpi_total"),  f"{n_resp}")
            rk2.metric(t("eval_kpi_useful"), f"{avg_useful:.1f} / 5.0")
            rk3.metric(t("eval_kpi_acc"),    f"{avg_accuracy:.1f} / 5.0")

            st.markdown("---")

            ec1, ec2 = st.columns(2)
            with ec1:
                useful_counts = eval_df["usefulness_rating"].dropna().value_counts().sort_index()
                fig_u = px.bar(
                    x=useful_counts.index.astype(str),
                    y=useful_counts.values,
                    labels={"x": t("eval_score_label"), "y": t("eval_responses")},
                    title=t("eval_chart_useful"),
                    color=useful_counts.values,
                    color_continuous_scale=["#EF4444", "#F59E0B", "#22C55E"],
                )
                fig_u.update_layout(
                    paper_bgcolor="#1E2D3D", plot_bgcolor="#1E2D3D",
                    font_color="#E2E8F0", height=280, showlegend=False,
                    coloraxis_showscale=False,
                    margin=dict(t=40, b=40, l=10, r=10),
                )
                st.plotly_chart(fig_u, use_container_width=True)

            with ec2:
                acc_counts = eval_df["accuracy_rating"].dropna().value_counts().sort_index()
                fig_a = px.bar(
                    x=acc_counts.index.astype(str),
                    y=acc_counts.values,
                    labels={"x": t("eval_score_label"), "y": t("eval_responses")},
                    title=t("eval_chart_acc"),
                    color=acc_counts.values,
                    color_continuous_scale=["#EF4444", "#F59E0B", "#22C55E"],
                )
                fig_a.update_layout(
                    paper_bgcolor="#1E2D3D", plot_bgcolor="#1E2D3D",
                    font_color="#E2E8F0", height=280, showlegend=False,
                    coloraxis_showscale=False,
                    margin=dict(t=40, b=40, l=10, r=10),
                )
                st.plotly_chart(fig_a, use_container_width=True)

            if "time_savings" in eval_df.columns and eval_df["time_savings"].notna().any():
                st.markdown(t("eval_time_title"))
                ts_counts = eval_df["time_savings"].value_counts()
                for option, count in ts_counts.items():
                    pct = count / n_resp * 100
                    st.markdown(
                        f"<div style='margin-bottom:6px'>"
                        f"<span style='color:#94A3B8;font-size:0.85rem'>{option}</span><br>"
                        f"<div style='background:#0D1B2A;border-radius:4px;height:8px;width:100%'>"
                        f"<div style='background:#3B82F6;border-radius:4px;height:8px;width:{int(pct)}%'></div></div>"
                        f"<span style='color:#64748B;font-size:0.78rem'>{count} {t('eval_response_lbl')} — {pct:.0f}%</span>"
                        f"</div>",
                        unsafe_allow_html=True,
                    )

            st.markdown("---")

            if "recommendation" in eval_df.columns and eval_df["recommendation"].notna().any():
                st.markdown(t("eval_rec_title"))
                rec_counts = eval_df["recommendation"].value_counts()
                fig_r = px.pie(
                    values=rec_counts.values,
                    names=rec_counts.index,
                    color_discrete_sequence=["#22C55E", "#3B82F6", "#F59E0B", "#EF4444"],
                    hole=0.45,
                )
                fig_r.update_layout(
                    paper_bgcolor="#1E2D3D",
                    font_color="#E2E8F0", height=260,
                    margin=dict(t=20, b=20, l=10, r=10),
                    legend=dict(font=dict(size=11)),
                )
                st.plotly_chart(fig_r, use_container_width=True)

            feat_reqs = eval_df["feature_requests"].dropna()
            feat_reqs = feat_reqs[feat_reqs.str.strip() != ""]
            if not feat_reqs.empty:
                st.markdown("---")
                st.subheader(t("eval_feat_title"))
                for req in feat_reqs:
                    st.markdown(
                        f"<div style='background:#1E2D3D;border:1px solid #2D3F50;border-left:3px solid #3B82F6;"
                        f"border-radius:8px;padding:0.7rem 1rem;margin-bottom:0.5rem;font-size:0.87rem;color:#CBD5E1'>"
                        f"{req}</div>",
                        unsafe_allow_html=True,
                    )

            comments = eval_df["general_comments"].dropna()
            comments = comments[comments.str.strip() != ""]
            if not comments.empty:
                st.markdown("---")
                st.subheader(t("eval_comm_title"))
                for comment in comments:
                    st.markdown(
                        f"<div style='background:#1E2D3D;border:1px solid #2D3F50;border-left:3px solid #F59E0B;"
                        f"border-radius:8px;padding:0.7rem 1rem;margin-bottom:0.5rem;font-size:0.87rem;color:#CBD5E1;font-style:italic'>"
                        f"\"{comment}\"</div>",
                        unsafe_allow_html=True,
                    )

            st.markdown("---")
            csv_eval = eval_df.to_csv(index=False).encode("utf-8")
            st.download_button(
                t("eval_export_btn"),
                csv_eval,
                f"platform_evaluation_{datetime.now().strftime('%Y%m%d')}.csv",
                "text/csv",
                use_container_width=False,
            )
